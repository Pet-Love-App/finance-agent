from __future__ import annotations

import errno
import datetime
import hashlib
import json
import os
import re
import signal
import sys
import textwrap
import urllib.error
import urllib.request
import zipfile
from urllib.parse import urlparse
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional, Set, Tuple
from collections import deque
import atexit
import copy
import uuid
import threading
import time

CURRENT_FILE = Path(__file__).resolve()


def _resolve_project_root() -> Path:
    env_root = os.getenv("AGENT_PROJECT_ROOT", "").strip()
    if env_root:
        candidate = Path(env_root).expanduser().resolve()
        if candidate.exists():
            return candidate

    for parent in (CURRENT_FILE, *CURRENT_FILE.parents):
        agent_dir = parent / "agent" / "__init__.py"
        data_dir = parent / "data"
        if agent_dir.exists() and data_dir.exists():
            return parent

    if len(CURRENT_FILE.parents) >= 3:
        return CURRENT_FILE.parents[2]
    return CURRENT_FILE.parent


# 将仓库根目录加入 sys.path，便于导入已有 reimbursement_agent 包
PROJECT_ROOT = _resolve_project_root()
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

SYSTEM_PROMPT = (
    "你是企业报销与办公任务助手。"
    "请使用自然、友好、专业的中文回答，先理解用户真实目标，再给出结论。"
    "默认采用“结论 + 关键步骤 + 注意事项”的结构，尽量短句表达，避免堆砌术语。"
    "当信息不足时，先给一个可执行的初步方案，再明确指出需要用户补充的关键信息。"
    "若涉及编辑文件但目标文件或改动范围不明确，先用一句话确认路径/文件名后再执行。"
    "若用户询问报销审计规则，请结合常见合规点和风险等级给出建议；"
    "若问题超出报销场景，也可进行通用问答并保持同样风格。"
)

DEFAULT_KB_PATH = PROJECT_ROOT / "data" / "kb" / "reimbursement_kb.json"

WORKSPACE_SKIP_DIRS = {
    ".git",
    ".idea",
    ".vscode",
    "node_modules",
    "dist",
    "build",
    "__pycache__",
    ".venv",
    "venv",
}

TEXT_EDIT_BLOCKED_SUFFIXES = {
    ".xlsx",
    ".xls",
    ".xlsm",
    ".xlsb",
    ".docx",
    ".doc",
    ".pptx",
    ".ppt",
    ".pdf",
    ".png",
    ".jpg",
    ".jpeg",
    ".gif",
    ".webp",
    ".bmp",
    ".zip",
    ".rar",
    ".7z",
}

XLSX_EDIT_SUFFIXES = {".xlsx", ".xlsm"}

DEFAULT_REIMBURSE_CATEGORY_KEYWORDS: Dict[str, Set[str]] = {
    "报销单": {"报销单", "报销申请", "报销", "reimburse"},
    "发票": {"发票", "invoice", "票据", "电子票"},
    "支付凭证": {"支付", "付款", "转账", "流水", "回单", "payment"},
    "费用明细": {"明细", "清单", "detail"},
    "活动说明": {"活动说明", "情况说明", "说明", "通知", "邮件", "mail"},
    "预算材料": {"预算", "budget"},
    "决算材料": {"决算", "final", "结项"},
    "签到材料": {"签到", "签名", "出席"},
}

DEFAULT_REQUIRED_CATEGORIES = ["报销单", "发票", "支付凭证", "费用明细"]

DEFAULT_MISSING_SUGGESTIONS: Dict[str, str] = {
    "报销单": "示例：报销单.xlsx / 报销申请表.docx",
    "发票": "示例：发票1.pdf / 电子发票.png",
    "支付凭证": "示例：支付回单.pdf / 转账截图.jpg",
    "费用明细": "示例：费用明细.xlsx / 报销清单.csv",
}


DEFAULT_MEMORY_PATH = PROJECT_ROOT / "data" / "memory" / "agent_memory.json"
DEFAULT_MEMORY_SHORT_TERM_LIMIT = 14
DEFAULT_MEMORY_RECENT_CONTEXT = 6
DEFAULT_MEMORY_LONG_TERM_LIMIT = 24
DEFAULT_MEMORY_SUMMARY_MAX_CHARS = 2400
DEFAULT_HISTORY_EXTRACT_THRESHOLD = 12

# In-memory cache + simple cross-process lock for safe writes
_MEMORY_CACHE_LOCK = threading.Lock()
_MEMORY_CACHE: Optional[Dict[str, Any]] = None
_MEMORY_DIRTY = False
_MEMORY_FLUSH_INTERVAL = 1.0  # seconds
# Event to signal the background flush thread to exit and perform a final flush
_MEMORY_FLUSH_STOP_EVENT = threading.Event()


def _acquire_path_lock(path: Path, timeout: float = 5.0):
    """Acquire an exclusive lock for ``path`` using a sibling lock file.

    The function repeatedly attempts to create a ``.lock`` file next to the
    target path using exclusive creation semantics. If successful, it writes
    the current process ID into the lock file and returns the information
    needed to release the lock later with ``_release_path_lock``.

    Args:
        path: The target file path to protect with a lock file.
        timeout: Maximum number of seconds to wait for the lock before
            raising an exception.

    Returns:
        A tuple ``(fd, lock_path)`` where ``fd`` is the open file descriptor
        for the created lock file and ``lock_path`` is the path of that lock
        file.

    Raises:
        TimeoutError: If the lock cannot be acquired within ``timeout``
            seconds.
        OSError: If creating or writing the lock file fails for a reason
            other than the lock file already existing.
    """
    lock_path = path.with_suffix(path.suffix + ".lock")
    start = time.time()
    while True:
        try:
            fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_RDWR, 0o600)
            os.write(fd, str(os.getpid()).encode())
            return fd, lock_path
        except FileExistsError:
            if time.time() - start > timeout:
                raise TimeoutError(f"Timeout acquiring lock for {path}")
            time.sleep(0.05)


def _release_path_lock(fd: int, lock_path: Path) -> None:
    try:
        os.close(fd)
    except Exception:
        pass
    try:
        if lock_path.exists():
            lock_path.unlink()
    except Exception:
        pass


def _save_memory_store_immediate(store: Dict[str, Any]) -> None:
    """Write store to disk immediately using an atomic temp file replace and lock."""
    path = _memory_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_suffix(path.suffix + ".tmp")
    fd = None
    lock_path = None
    try:
        fd, lock_path = _acquire_path_lock(path)
        temp_path.write_text(json.dumps(store, ensure_ascii=False, indent=2), encoding="utf-8")
        temp_path.replace(path)
    finally:
        if fd is not None and lock_path is not None:
            _release_path_lock(fd, lock_path)


def _memory_flush_daemon() -> None:
    global _MEMORY_DIRTY
    # Periodically wake up to flush dirty cache. Exit when stop event is set.
    try:
        while not _MEMORY_FLUSH_STOP_EVENT.wait(_MEMORY_FLUSH_INTERVAL):
            try:
                snapshot = None
                with _MEMORY_CACHE_LOCK:
                    if _MEMORY_DIRTY and _MEMORY_CACHE is not None:
                        snapshot = copy.deepcopy(_MEMORY_CACHE)
                if snapshot is None:
                    continue
                _save_memory_store_immediate(snapshot)
                with _MEMORY_CACHE_LOCK:
                    if _MEMORY_DIRTY and _MEMORY_CACHE == snapshot:
                        _MEMORY_DIRTY = False
            except Exception:
                # ignore flush errors; next cycle will retry
                pass
    finally:
        # On thread exit, attempt one final immediate flush to reduce data loss.
        try:
            snapshot = None
            with _MEMORY_CACHE_LOCK:
                if _MEMORY_DIRTY and _MEMORY_CACHE is not None:
                    snapshot = copy.deepcopy(_MEMORY_CACHE)
            if snapshot is not None:
                try:
                    _save_memory_store_immediate(snapshot)
                    with _MEMORY_CACHE_LOCK:
                        if _MEMORY_DIRTY and _MEMORY_CACHE == snapshot:
                            _MEMORY_DIRTY = False
                except Exception:
                    # last-resort: ignore
                    pass
        except Exception:
            pass


# background flush thread handle and lock. Thread will be started lazily
# when memory operations require it (avoids issues during import/test).
_MEMORY_FLUSH_THREAD: Optional[threading.Thread] = None
_MEMORY_FLUSH_THREAD_LOCK = threading.Lock()


def _ensure_memory_flush_thread_started() -> None:
    """Start the memory flush thread lazily. Safe to call multiple times.

    Use this from code paths that mutate memory (e.g. `_save_memory_store`) so
    the thread isn't created at import time and test environments can control
    lifecycle.
    """
    global _MEMORY_FLUSH_THREAD
    try:
        with _MEMORY_FLUSH_THREAD_LOCK:
            if _MEMORY_FLUSH_THREAD is None or not _MEMORY_FLUSH_THREAD.is_alive():
                _MEMORY_FLUSH_THREAD = threading.Thread(target=_memory_flush_daemon, daemon=False)
                _MEMORY_FLUSH_THREAD.start()
    except Exception:
        # If starting the thread fails, fall back silently; writes will still
        # be persisted by immediate-flush code paths.
        pass


def start_memory_flush_thread() -> None:
    """Public helper to start the background flush thread."""
    _ensure_memory_flush_thread_started()


def stop_memory_flush_thread(timeout: float = 5.0) -> None:
    """Public helper to stop the background flush thread and flush data."""
    _shutdown_memory_flush(timeout=timeout)


def _shutdown_memory_flush(timeout: float = 5.0) -> None:
    """Signal the memory flush thread to stop, wait for it, and perform a final flush.

    This is safe to call multiple times. It will set the stop event, join the
    background thread (with a timeout), and attempt an immediate flush if any
    dirty data remains.
    """
    try:
        _MEMORY_FLUSH_STOP_EVENT.set()
    except Exception:
        pass
    try:
        if _MEMORY_FLUSH_THREAD is not None and _MEMORY_FLUSH_THREAD.is_alive():
            try:
                _MEMORY_FLUSH_THREAD.join(timeout)
            except Exception:
                pass
    except Exception:
        pass

    # final immediate flush if still dirty
    try:
        with _MEMORY_CACHE_LOCK:
            if _MEMORY_DIRTY and _MEMORY_CACHE is not None:
                try:
                    _save_memory_store_immediate(_MEMORY_CACHE)
                    _MEMORY_DIRTY = False
                except Exception:
                    pass
    except Exception:
        pass


# Ensure graceful shutdown on normal interpreter exit
try:
    atexit.register(_shutdown_memory_flush)
except Exception:
    pass


def _memory_enabled(payload: Dict[str, Any]) -> bool:
    raw = payload.get("memory_enabled", True)
    if isinstance(raw, bool):
        return raw
    text = str(raw).strip().lower()
    return text not in {"0", "false", "off", "no"}


def _env_bool(key: str, default: bool = False) -> bool:
    raw = os.getenv(key, "").strip()
    if not raw:
        return default
    return str(raw).strip().lower() not in {"0", "false", "off", "no"}


def _memory_path() -> Path:
    raw = os.getenv("AGENT_MEMORY_PATH", "").strip()
    if raw:
        return Path(raw).expanduser().resolve()
    return DEFAULT_MEMORY_PATH


def _memory_session_key(payload: Dict[str, Any]) -> str:
    scope = str(payload.get("memory_scope", "workspace")).strip().lower() or "workspace"
    workspace = str(payload.get("workspace_dir", "")).strip()
    base = workspace if workspace else "default"
    digest = hashlib.sha1(base.encode("utf-8")).hexdigest()[:16]
    return f"{scope}:{digest}"


def _load_memory_store() -> Dict[str, Any]:
    global _MEMORY_CACHE
    path = _memory_path()
    # if cached in-memory, return a shallow copy to callers
    with _MEMORY_CACHE_LOCK:
        if _MEMORY_CACHE is not None:
            try:
                # Prefer a true deepcopy; fall back to a shallow dict copy if deepcopy fails.
                return copy.deepcopy(_MEMORY_CACHE)
            except Exception:
                try:
                    return dict(_MEMORY_CACHE)
                except Exception:
                    # fallback to reading from disk
                    pass
    if not path.exists():
        return {"version": 1, "sessions": {}}
    fd = None
    lock_path = None
    try:
        fd, lock_path = _acquire_path_lock(path)
        try:
            raw = path.read_text(encoding="utf-8")
            parsed = json.loads(raw)
        except Exception:
            parsed = {"version": 1, "sessions": {}}
        if not isinstance(parsed, dict):
            parsed = {"version": 1, "sessions": {}}
        sessions = parsed.get("sessions")
        if not isinstance(sessions, dict):
            parsed["sessions"] = {}
        parsed.setdefault("version", 1)
        with _MEMORY_CACHE_LOCK:
            _MEMORY_CACHE = parsed
        return parsed
    finally:
        if fd is not None and lock_path is not None:
            _release_path_lock(fd, lock_path)


def _save_memory_store(store: Dict[str, Any]) -> None:
    # update in-memory cache and mark dirty; background thread will flush
    global _MEMORY_CACHE, _MEMORY_DIRTY
    with _MEMORY_CACHE_LOCK:
        _MEMORY_CACHE = store
        _MEMORY_DIRTY = True

    # Ensure the background flush thread is running so the dirty cache will be
    # flushed. Lazy start avoids creating threads at import time (helps tests).
    try:
        _ensure_memory_flush_thread_started()
    except Exception:
        pass

    # Optional immediate flush for short-lived child processes
    try:
        if _env_bool("AGENT_MEMORY_IMMEDIATE_FLUSH", False):
            try:
                _save_memory_store_immediate(store)
                with _MEMORY_CACHE_LOCK:
                    _MEMORY_DIRTY = False
            except Exception:
                # fall back to background flush
                pass
    except Exception:
        pass


def _get_or_create_memory_session(store: Dict[str, Any], session_key: str) -> Dict[str, Any]:
    sessions = store.setdefault("sessions", {})
    if not isinstance(sessions, dict):
        sessions = {}
        store["sessions"] = sessions
    existing = sessions.get(session_key)
    if isinstance(existing, dict):
        existing.setdefault("short_term", [])
        existing.setdefault("long_term", [])
        existing.setdefault("rolling_summary", "")
        existing.setdefault("profile", {})
        return existing
    now = datetime.datetime.now().isoformat()
    session = {
        "updated_at": now,
        "short_term": [],
        "long_term": [],
        "rolling_summary": "",
        "profile": {},
    }
    sessions[session_key] = session
    return session


def _reset_memory_session(payload: Dict[str, Any]) -> None:
    if not _memory_enabled(payload):
        return
    store = _load_memory_store()
    session_key = _memory_session_key(payload)
    sessions = store.setdefault("sessions", {})
    if isinstance(sessions, dict):
        sessions.pop(session_key, None)
    # Always update the in-memory cache for this process first, then attempt
    # an immediate flush so reset operations become durable on disk promptly.
    _save_memory_store(store)
    try:
        _save_memory_store_immediate(store)
        with _MEMORY_CACHE_LOCK:
            _MEMORY_DIRTY = False
    except Exception:
        # Cache is already updated and marked dirty; background flush will
        # persist it if the immediate write fails.
        pass


def _summarize_messages(messages: List[Dict[str, str]], *, max_chars: int = 900) -> str:
    lines: List[str] = []
    for item in messages:
        role = "用户" if item.get("role") == "user" else "助手"
        content = str(item.get("content", "")).strip().replace("\n", " ")
        if not content:
            continue
        lines.append(f"- {role}: {content[:140]}")
    text = "\n".join(lines)
    return text[:max_chars]


def _compute_importance(text: str, *, role: str = "user") -> float:
    """Compute a heuristic importance score used for memory promotion decisions.

    The score estimates how likely a message is to contain information worth
    retaining beyond the short-term conversation buffer. This is a heuristic,
    additive scoring algorithm:

    * Empty input returns ``0.0`` immediately.
    * Non-empty input starts with a base score of ``0.2``.
    * Add ``1.2`` if the text contains explicit imperative or memory-related
      keywords such as ``记住``、``重要``、``必须``、``务必`` or ``一定``.
    * Add ``0.8`` if the original text matches a personal profile pattern such
      as ``我叫...``.
    * Add ``0.4`` if the text contains dates or numeric facts, including forms
      like ``YYYY-MM-DD``, ``MM月``, ``DD日``, ``N年`` or other digits.
    * Add ``0.3`` for short factual statements (under 120 characters) that
      include cues such as ``是``、``来自`` or ``公司``.

    The final value is capped at ``2.5`` and rounded to three decimal places,
    so the effective return range is ``0.0`` to ``2.5`` inclusive.

    Args:
        text: Message content to evaluate.
        role: Logical speaker role for the message. It is currently accepted for
            interface consistency and future heuristic extensions, but it does
            not affect the score in the current implementation.

    Returns:
        A floating-point importance score in the range ``0.0`` to ``2.5``,
        where higher values indicate that the message is more likely to be
        promoted into longer-term memory.
    """
    if not text:
        return 0.0
    score = 0.2
    t = text.lower()
    # explicit imperative keywords
    if any(k in t for k in ("记住", "重要", "必须", "务必", "一定")):
        score += 1.2
    # personal profile signals
    if re.search(r"我叫[^\s，。,；;]+", text):
        score += 0.8
    # presence of dates/numbers suggests factual content
    if re.search(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}月|\d{1,2}日|\d+年|\d+", text):
        score += 0.4
    # short explicit facts
    if len(text) < 120 and ("是" in t or "来自" in t or "公司" in t):
        score += 0.3
    # cap
    return round(min(score, 2.5), 3)


def _extract_memory_facts(message: str) -> List[Dict[str, str]]:
    text = str(message or "").strip()
    if not text:
        return []
    facts: List[Dict[str, str]] = []
    patterns: List[Tuple[str, str]] = [
        (r"(?:请|帮我)?记住[:：]?\s*(.+)", "explicit"),
        (r"我叫([^\s，。,；;]+)", "profile"),
        (r"请用([^。！!\n]+)", "preference"),
        (r"(?:以后|之后).{0,20}(?:不要|别)([^。！!\n]+)", "constraint"),
        (r"我的(?:偏好|习惯|风格)是([^。！!\n]+)", "preference"),
    ]
    for pattern, kind in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if not match:
            continue
        value = str(match.group(1)).strip("：:，,。；; ")
        if not value:
            continue
        facts.append({"type": kind, "fact": value[:240]})
    if ("不要" in text or "必须" in text or "务必" in text) and len(text) <= 180:
        facts.append({"type": "constraint", "fact": text})
    unique: List[Dict[str, str]] = []
    seen: Set[str] = set()
    for item in facts:
        fact = str(item.get("fact", "")).strip()
        if not fact or fact in seen:
            continue
        seen.add(fact)
        unique.append(item)
    return unique


def _merge_memory_profile(session: Dict[str, Any], memory_profile: Any) -> None:
    if not isinstance(memory_profile, dict):
        return
    profile = session.setdefault("profile", {})
    if not isinstance(profile, dict):
        profile = {}
        session["profile"] = profile
    for key, value in memory_profile.items():
        normalized_key = str(key).strip()
        if not normalized_key:
            continue
        profile[normalized_key] = str(value).strip()[:200]

    # Build a temporary working deque that is larger than the final short-term limit
    # so we can score, deduplicate, and promote entries before trimming back down.
    # `short_limit * 2` keeps roughly one extra window of candidates available during
    # that selection pass, while `short_limit + 4` guarantees a small fixed cushion
    # when the configured short-term limit is near its minimum.
def _remember_turn(payload: Dict[str, Any], user_message: str, assistant_reply: str) -> None:
    if not _memory_enabled(payload):
        return
    store = _load_memory_store()
    session_key = _memory_session_key(payload)
    session = _get_or_create_memory_session(store, session_key)

    _merge_memory_profile(session, payload.get("memory_profile"))

    # handle short_term as a deque with importance scoring and promotion
    raw_short = session.setdefault("short_term", [])
    if not isinstance(raw_short, list):
        raw_short = []
        session["short_term"] = raw_short

    short_limit = _safe_int_env("AGENT_MEMORY_SHORT_TERM_LIMIT", DEFAULT_MEMORY_SHORT_TERM_LIMIT, min_value=6)
    summary_max_chars = _safe_int_env("AGENT_MEMORY_SUMMARY_MAX_CHARS", DEFAULT_MEMORY_SUMMARY_MAX_CHARS, min_value=800)

    user_text = str(user_message or "").strip()
    reply_text = str(assistant_reply or "").strip()

    # build working deque (allow some overflow to choose important messages)
    work_max = max(short_limit * 2, short_limit + 4)
    dq = deque(raw_short, maxlen=work_max)

    now_ts = datetime.datetime.now().isoformat()
    # Also ingest recent user messages from the frontend-provided history (context window)
    try:
        history_items = payload.get("history", []) if isinstance(payload, dict) else []
        if isinstance(history_items, list) and history_items:
            # take up to last `short_limit` user messages from history (older -> newer)
            user_hist = [str(it.get("content", "")).strip() for it in history_items if isinstance(it, dict) and it.get("role") == "user"]
            if user_hist:
                # avoid duplicates already in short_term
                existing_contents = {str(it.get("content", "")).strip() for it in dq if isinstance(it, dict)}
                for txt in user_hist[-short_limit:]:
                    if not txt:
                        continue
                    if txt in existing_contents:
                        continue
                    entry = {
                        "id": uuid.uuid4().hex,
                        "role": "user",
                        "content": txt[:2000],
                        "ts": now_ts,
                        "importance": _compute_importance(txt, role="user"),
                    }
                    dq.append(entry)
                    existing_contents.add(txt)
    except Exception:
        # be conservative: don't let history ingestion break memory flow
        pass
    if user_text:
        entry = {
            "id": uuid.uuid4().hex,
            "role": "user",
            "content": user_text[:2000],
            "ts": now_ts,
            "importance": _compute_importance(user_text, role="user"),
        }
        dq.append(entry)
    if reply_text:
        entry = {
            "id": uuid.uuid4().hex,
            "role": "assistant",
            "content": reply_text[:2200],
            "ts": now_ts,
            "importance": _compute_importance(reply_text, role="assistant"),
        }
        dq.append(entry)

    rolling_summary = str(session.get("rolling_summary", "") or "")

    # trim using priority: keep recent + top-important from older
    if len(dq) > short_limit:
        keep_recent = max(short_limit // 2, 4)
        list_all = list(dq)

        # Normalize ids for legacy persisted entries and avoid duplicate ids so
        # trimming/overflow logic can reliably identify individual messages.
        seen_ids: Set[str] = set()
        for item in list_all:
            item_id = item.get("id")
            if not item_id or item_id in seen_ids:
                item_id = uuid.uuid4().hex
                item["id"] = item_id
            seen_ids.add(item_id)

        recent = list_all[-keep_recent:]
        older = list_all[:-keep_recent]
        need = max(short_limit - keep_recent, 0)
        # select top important from older
        selected = sorted(older, key=lambda x: float(x.get("importance", 0.0)), reverse=True)[:need]
        kept_ids = {it["id"] for it in recent}
        kept_ids.update(it["id"] for it in selected)
        # Preserve original chronological order when combining recent and selected.
        final_short = [it for it in list_all if it["id"] in kept_ids]
        # compute overflow
        overflow = [it for it in list_all if it["id"] not in kept_ids]
        # update session short_term as list for persistence
        session["short_term"] = final_short
        # summarize overflow into rolling_summary
        overflow_summary = _summarize_messages(overflow, max_chars=1000)
        if overflow_summary:
            rolling_summary = (rolling_summary + "\n" + overflow_summary).strip() if rolling_summary else overflow_summary
            session["rolling_summary"] = rolling_summary[:summary_max_chars]
    else:
        # no trimming needed; persist current deque
        session["short_term"] = list(dq)

    long_term = session.setdefault("long_term", [])
    if not isinstance(long_term, list):
        long_term = []
        session["long_term"] = long_term

    # Promote high-importance short-term messages to long_term if they contain extractable facts
    promote_threshold = 1.0
    now = datetime.datetime.now().isoformat()
    existing_facts = {str(item.get("fact", "")) for item in long_term if isinstance(item, dict)}
    try:
        short_entries = session.get("short_term", []) or []
        for entry in short_entries:
            try:
                if not isinstance(entry, dict):
                    continue
                if entry.get("role") != "user":
                    continue
                importance = float(entry.get("importance", 0.0))
                if importance < promote_threshold:
                    continue
                facts = _extract_memory_facts(str(entry.get("content", "")))
                for item in facts:
                    fact = str(item.get("fact", "")).strip()
                    if not fact or fact in existing_facts:
                        continue
                    long_term.append({
                        "type": str(item.get("type", "fact")),
                        "fact": fact,
                        "source": "auto_promote",
                        "updated_at": now,
                    })
                    existing_facts.add(fact)
            except Exception:
                continue
    except Exception:
        pass

    extracted = _extract_memory_facts(user_text)
    if extracted:
        now = datetime.datetime.now().isoformat()
        existing_facts = {str(item.get("fact", "")) for item in long_term if isinstance(item, dict)}
        for item in extracted:
            fact = str(item.get("fact", "")).strip()
            if not fact or fact in existing_facts:
                continue
            long_term.append(
                {
                    "type": str(item.get("type", "fact")),
                    "fact": fact,
                    "source": "user_message",
                    "updated_at": now,
                }
            )
            existing_facts.add(fact)
    long_limit = _safe_int_env("AGENT_MEMORY_LONG_TERM_LIMIT", DEFAULT_MEMORY_LONG_TERM_LIMIT, min_value=6)
    if len(long_term) > long_limit:
        session["long_term"] = long_term[-long_limit:]

    session["updated_at"] = datetime.datetime.now().isoformat()
    # If frontend provided a long history and it reaches threshold, run extraction+immediate flush
    try:
        history_items = payload.get("history", []) if isinstance(payload, dict) else []
        history_count = len(history_items) if isinstance(history_items, list) else 0
        extract_threshold = _safe_int_env("AGENT_MEMORY_EXTRACT_HISTORY_THRESHOLD", DEFAULT_HISTORY_EXTRACT_THRESHOLD, min_value=1)
        if history_count >= extract_threshold:
            try:
                existing_facts = {str(item.get("fact", "")) for item in long_term if isinstance(item, dict)}
                now_iso = datetime.datetime.now().isoformat()
                # examine last `extract_threshold` user messages
                user_hist = [str(it.get("content", "")).strip() for it in history_items if isinstance(it, dict) and it.get("role") == "user"]
                for txt in user_hist[-extract_threshold:]:
                    if not txt:
                        continue
                    facts = _extract_memory_facts(txt)
                    for item in facts:
                        fact = str(item.get("fact", "")).strip()
                        if not fact or fact in existing_facts:
                            continue
                        # promote if explicit or importance meets threshold
                        score = _compute_importance(txt, role="user")
                        promote_threshold = _safe_int_env("AGENT_MEMORY_PROMOTE_THRESHOLD", 1, min_value=0)
                        if item.get("type") == "explicit" or float(score) >= float(promote_threshold):
                            long_term.append({
                                "type": str(item.get("type", "fact")),
                                "fact": fact,
                                "source": "auto_from_history",
                                "updated_at": now_iso,
                            })
                            existing_facts.add(fact)
                # trim long_term
                long_limit = _safe_int_env("AGENT_MEMORY_LONG_TERM_LIMIT", DEFAULT_MEMORY_LONG_TERM_LIMIT, min_value=6)
                if len(long_term) > long_limit:
                    session["long_term"] = long_term[-long_limit:]
                # immediate flush to disk so short-lived processes persist
                try:
                    _save_memory_store_immediate(store)
                    with _MEMORY_CACHE_LOCK:
                        _MEMORY_DIRTY = False
                except Exception:
                    _save_memory_store(store)
            except Exception:
                # ignore extraction errors
                pass
    except Exception:
        pass
    _save_memory_store(store)


def _memory_context(payload: Dict[str, Any]) -> str:
    if not _memory_enabled(payload):
        return ""
    store = _load_memory_store()
    session = _get_or_create_memory_session(store, _memory_session_key(payload))

    chunks: List[str] = []
    rolling_summary = str(session.get("rolling_summary", "") or "").strip()
    if rolling_summary:
        chunks.append(f"历史摘要:\n{rolling_summary}")

    recent_limit = _safe_int_env("AGENT_MEMORY_RECENT_CONTEXT", DEFAULT_MEMORY_RECENT_CONTEXT, min_value=2)
    short_term = session.get("short_term", [])
    if isinstance(short_term, list) and short_term:
        rows: List[str] = []
        for item in short_term[-recent_limit:]:
            if not isinstance(item, dict):
                continue
            role = "用户" if item.get("role") == "user" else "助手"
            content = str(item.get("content", "")).strip()
            if content:
                rows.append(f"- {role}: {content[:180]}")
        if rows:
            chunks.append("最近对话记忆:\n" + "\n".join(rows))

    long_term = session.get("long_term", [])
    if isinstance(long_term, list) and long_term:
        facts: List[str] = []
        for item in long_term[-8:]:
            if not isinstance(item, dict):
                continue
            fact = str(item.get("fact", "")).strip()
            if fact:
                facts.append(f"- {fact[:180]}")
        if facts:
            chunks.append("长期记忆:\n" + "\n".join(facts))

    profile = session.get("profile", {})
    if isinstance(profile, dict) and profile:
        pairs = [f"{str(key).strip()}={str(value).strip()}" for key, value in profile.items() if str(key).strip()]
        if pairs:
            chunks.append("用户画像:\n" + "\n".join(f"- {item[:180]}" for item in pairs[:12]))
    return "\n\n".join(chunks).strip()


def _merge_context_blocks(*blocks: str) -> str:
    cleaned = [str(item).strip() for item in blocks if str(item).strip()]
    return "\n\n".join(cleaned)


def _referenced_files_context(payload: Dict[str, Any]) -> str:
    blocks: List[str] = []
    raw_files = payload.get("referenced_files", [])
    if isinstance(raw_files, list) and raw_files:
        files = [str(item).strip() for item in raw_files if str(item).strip()]
        if files:
            blocks.append("用户指定参考文件:\n" + "\n".join(f"- {item}" for item in files[:20]))

    raw_context = str(payload.get("referenced_file_context", "")).strip()
    if raw_context:
        blocks.append("用户引用文件内容:\n" + raw_context[:24000])

    return _merge_context_blocks(*blocks)


def _referenced_file_paths(payload: Dict[str, Any]) -> List[str]:
    raw_files = payload.get("referenced_files", [])
    if not isinstance(raw_files, list):
        return []
    return [str(item).strip() for item in raw_files if str(item).strip()]


def _resolve_message_with_referenced_file(message: str, payload: Dict[str, Any]) -> str:
    text = str(message or "").strip()
    if not text:
        return text
    referenced_files = _referenced_file_paths(payload)
    if len(referenced_files) != 1:
        return text

    target = referenced_files[0]
    lowered = text.lower()
    has_pronoun = any(token in text for token in ("这个文件", "当前文件", "该文件", "这个表", "这个excel"))
    has_target = target in text or target.lower() in lowered
    has_file_ext = bool(
        re.search(
            r"\.(py|ts|tsx|js|jsx|json|md|txt|yaml|yml|toml|ini|csv|xlsx|xlsm|docx|pdf)\b",
            text,
            flags=re.IGNORECASE,
        )
    )
    if has_target or has_file_ext or not has_pronoun:
        return text

    return f"{text}\n\n用户指代的目标文件是：{target}"


def _build_direct_plan_from_single_reference(message: str, payload: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    text = str(message or "").strip()
    if not text:
        return None
    referenced_files = _referenced_file_paths(payload)
    if len(referenced_files) != 1:
        return None

    target = referenced_files[0]
    lower = text.lower()
    has_data_intent = any(token in text for token in ("测试数据", "追加", "加入", "新增")) or "append" in lower
    if not has_data_intent:
        return None

    if not target.lower().endswith((".xlsx", ".xlsm")):
        return None

    count_match = re.search(r"(\d+)\s*条", text)
    count = int(count_match.group(1)) if count_match else 5
    count = max(1, min(count, 2000))
    sheet_match = re.search(r"\b(Sheet\d+)\b", text, flags=re.IGNORECASE)
    requested_sheet = sheet_match.group(1) if sheet_match else "Sheet1"
    rows = [{"学号": 1000 + i + 1, "姓名": f"测试{i + 1}"} for i in range(count)]
    return {
        "reply": f"已按你引用的文件执行：向 {target} 的 {requested_sheet} 追加 {count} 条测试数据。",
        "actions": [
            {
                "action": "xlsx_edit",
                "path": target,
                "sheet": requested_sheet,
                "append_dict_rows": rows,
            }
        ],
    }


def _safe_workspace_root(payload: Dict[str, Any]) -> Optional[Path]:
    workspace_raw = str(payload.get("workspace_dir", "")).strip()
    if not workspace_raw:
        return None
    try:
        root = Path(workspace_raw).expanduser().resolve()
    except Exception:
        return None
    if not root.exists() or not root.is_dir():
        return None
    return root


def _safe_workspace_target(root: Path, relative_path: str) -> Path:
    rel = str(relative_path or "").strip().replace("\\", "/")
    if not rel:
        raise ValueError("路径不能为空")
    target = (root / rel).resolve()
    try:
        target.relative_to(root)
    except ValueError as exc:
        raise ValueError("禁止访问目录外路径") from exc
    return target


def _workspace_tree_text(root: Path, *, max_files: Optional[int] = None) -> str:
    if max_files is None:
        max_files = _safe_int_env("AGENT_WORKSPACE_TREE_MAX_FILES", 5000, min_value=200)
    rows: List[str] = []
    count = 0
    for current, dirs, files in os.walk(root):
        dirs[:] = [d for d in dirs if d not in WORKSPACE_SKIP_DIRS and not d.startswith(".")]
        rel_dir = Path(current).resolve().relative_to(root)
        rel_prefix = "" if str(rel_dir) == "." else str(rel_dir).replace("\\", "/") + "/"
        for name in sorted(files):
            if name.startswith("."):
                continue
            rows.append(rel_prefix + name)
            count += 1
            if count >= max_files:
                rows.append("... (更多文件已省略)")
                return "\n".join(rows)
    return "\n".join(rows) if rows else "(空目录)"


def _workspace_all_files(root: Path, *, max_files: int = 5000) -> List[Path]:
    files: List[Path] = []
    for current, dirs, names in os.walk(root):
        dirs[:] = [d for d in dirs if d not in WORKSPACE_SKIP_DIRS and not d.startswith(".")]
        for name in sorted(names):
            if name.startswith("."):
                continue
            path = (Path(current) / name).resolve()
            try:
                path.relative_to(root)
            except ValueError:
                continue
            files.append(path)
            if len(files) >= max_files:
                return files
    return files


def _match_keywords(text: str, keywords: Set[str]) -> bool:
    lowered = text.lower()
    return any(key.lower() in lowered for key in keywords)


def _parse_reimburse_package_options(raw_options: Any) -> Tuple[Dict[str, Set[str]], List[str], Dict[str, str], bool]:
    category_keywords: Dict[str, Set[str]] = {
        key: set(values) for key, values in DEFAULT_REIMBURSE_CATEGORY_KEYWORDS.items()
    }
    required_categories = list(DEFAULT_REQUIRED_CATEGORIES)
    suggestions = dict(DEFAULT_MISSING_SUGGESTIONS)
    include_uncategorized = True

    if not isinstance(raw_options, dict):
        return category_keywords, required_categories, suggestions, include_uncategorized

    custom_keywords = raw_options.get("category_keywords")
    if isinstance(custom_keywords, dict):
        for category, values in custom_keywords.items():
            key = str(category).strip()
            if not key:
                continue
            if isinstance(values, list):
                words = {str(item).strip() for item in values if str(item).strip()}
                if words:
                    category_keywords[key] = words

    custom_required = raw_options.get("required_categories")
    if isinstance(custom_required, list):
        normalized_required = [str(item).strip() for item in custom_required if str(item).strip()]
        if normalized_required:
            required_categories = normalized_required
            for category in required_categories:
                category_keywords.setdefault(category, {category})

    custom_suggestions = raw_options.get("missing_suggestions")
    if isinstance(custom_suggestions, dict):
        for category, tip in custom_suggestions.items():
            key = str(category).strip()
            if not key:
                continue
            tip_text = str(tip).strip()
            if tip_text:
                suggestions[key] = tip_text

    include_uncategorized = bool(raw_options.get("include_uncategorized", True))
    return category_keywords, required_categories, suggestions, include_uncategorized


def _workspace_prepare_reimbursement_package(
    root: Path,
    package_name: Optional[str] = None,
    options: Optional[Dict[str, Any]] = None,
) -> str:
    all_files = _workspace_all_files(root)
    if not all_files:
        raise ValueError("目录为空，未找到可打包的材料。")

    category_keywords, required_categories, suggestion_map, include_uncategorized = _parse_reimburse_package_options(
        options or {}
    )

    category_files: Dict[str, List[Path]] = {key: [] for key in category_keywords.keys()}
    uncategorized: List[Path] = []

    for file_path in all_files:
        filename = file_path.name
        rel = str(file_path.relative_to(root)).replace("\\", "/")

        # 跳过历史打包结果，避免把 zip 包再次打进新 zip。
        if file_path.suffix.lower() == ".zip":
            continue

        matched = False
        for category, keywords in category_keywords.items():
            if _match_keywords(filename, keywords) or _match_keywords(rel, keywords):
                category_files[category].append(file_path)
                matched = True
        if not matched:
            uncategorized.append(file_path)

    missing = [name for name in required_categories if not category_files.get(name)]
    if missing:
        details = "\n".join(f"- 缺少：{name}（{suggestion_map.get(name, '请补充对应材料')}）" for name in missing)
        raise ValueError(f"检测到材料不完整，请先补齐后再打包：\n{details}")

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    raw_name = str(package_name or f"reimbursement_package_{timestamp}.zip").strip()
    if not raw_name.lower().endswith(".zip"):
        raw_name += ".zip"

    output_zip = (root / raw_name).resolve()
    try:
        output_zip.relative_to(root)
    except ValueError as exc:
        raise ValueError("压缩包名称非法，请仅提供文件名，不要包含目录穿越路径。") from exc

    with zipfile.ZipFile(output_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for category, files in category_files.items():
            for file_path in files:
                arcname = f"{category}/{file_path.name}"
                zf.write(file_path, arcname=arcname)
        if include_uncategorized:
            for file_path in uncategorized:
                arcname = f"其他材料/{file_path.name}"
                zf.write(file_path, arcname=arcname)

    total_count = sum(len(files) for files in category_files.values()) + (len(uncategorized) if include_uncategorized else 0)
    summary_items = [f"{name} {len(category_files.get(name, []))} 份" for name in required_categories]
    return (
        f"已生成压缩包：{output_zip.name}（共 {total_count} 个文件）\n"
        f"分类统计：{', '.join(summary_items)}"
    )


def _workspace_read(root: Path, relative_path: str) -> str:
    target = _safe_workspace_target(root, relative_path)
    if not target.exists() or not target.is_file():
        raise FileNotFoundError(f"文件不存在: {relative_path}")
    if target.suffix.lower() in TEXT_EDIT_BLOCKED_SUFFIXES:
        raise ValueError(f"该文件类型不支持文本读取: {target.name}")
    return target.read_text(encoding="utf-8")


def _workspace_write(root: Path, relative_path: str, content: str) -> None:
    target = _safe_workspace_target(root, relative_path)
    if target.suffix.lower() in TEXT_EDIT_BLOCKED_SUFFIXES:
        raise ValueError(f"该文件类型不支持文本写入: {target.name}")
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(content, encoding="utf-8")


def _workspace_append(root: Path, relative_path: str, content: str) -> None:
    target = _safe_workspace_target(root, relative_path)
    if target.suffix.lower() in TEXT_EDIT_BLOCKED_SUFFIXES:
        raise ValueError(f"该文件类型不支持文本追加: {target.name}")
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("a", encoding="utf-8") as f:
        f.write(content)


def _workspace_replace(root: Path, relative_path: str, old: str, new: str) -> int:
    target = _safe_workspace_target(root, relative_path)
    if target.suffix.lower() in TEXT_EDIT_BLOCKED_SUFFIXES:
        raise ValueError(f"该文件类型不支持文本替换: {target.name}")
    source = _workspace_read(root, relative_path)
    if old not in source:
        return 0
    updated = source.replace(old, new)
    _workspace_write(root, relative_path, updated)
    return source.count(old)


def _workspace_xlsx_edit(
    root: Path,
    relative_path: str,
    *,
    sheet: Optional[str],
    set_cells: Any,
    append_rows: Any,
    append_dict_rows: Any,
) -> str:
    target = _safe_workspace_target(root, relative_path)
    if target.suffix.lower() not in XLSX_EDIT_SUFFIXES:
        raise ValueError("xlsx_edit 仅支持 .xlsx 或 .xlsm 文件")

    try:
        import openpyxl  # type: ignore
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法执行 Excel 编辑。") from exc

    recovery_note = ""
    if target.exists():
        try:
            workbook = openpyxl.load_workbook(target)
        except Exception as exc:
            raw = str(exc)
            if "File is not a zip file" in raw:
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_path = target.with_name(f"{target.stem}.corrupt_{timestamp}{target.suffix}.bak")
                try:
                    target.replace(backup_path)
                except Exception as move_exc:
                    raise RuntimeError(
                        "检测到 Excel 文件已损坏，且备份失败。请关闭占用程序后重试，或手动备份该文件。"
                    ) from move_exc
                workbook = openpyxl.Workbook()
                recovery_note = f"检测到原文件损坏，已备份为 {backup_path.name} 并重建新文件。"
            else:
                raise RuntimeError(f"无法读取 Excel 文件：{raw}") from exc
    else:
        workbook = openpyxl.Workbook()
    sheet_name = (sheet or "").strip()
    if sheet_name:
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
        else:
            default_ws = workbook[workbook.sheetnames[0]]
            default_is_blank = (
                len(workbook.sheetnames) == 1
                and default_ws.max_row <= 1
                and default_ws.max_column <= 1
                and default_ws.cell(1, 1).value in (None, "")
            )
            if default_is_blank:
                default_ws.title = sheet_name
                ws = default_ws
            else:
                ws = workbook.create_sheet(sheet_name)
    else:
        ws = workbook[workbook.sheetnames[0]]

    set_count = 0
    append_count = 0
    append_dict_count = 0

    if isinstance(set_cells, list):
        for item in set_cells[:400]:
            if not isinstance(item, dict):
                continue
            cell_ref = str(item.get("cell", "")).strip().upper()
            if not cell_ref:
                continue
            ws[cell_ref] = item.get("value")
            set_count += 1

    if isinstance(append_rows, list):
        for row in append_rows[:1000]:
            if isinstance(row, list):
                ws.append(row)
                append_count += 1
            elif isinstance(row, tuple):
                ws.append(list(row))
                append_count += 1

    if isinstance(append_dict_rows, list):
        dict_rows = [item for item in append_dict_rows[:1000] if isinstance(item, dict)]
        if dict_rows:
            first_row_values = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
            header_map: Dict[str, int] = {}
            for index, value in enumerate(first_row_values, start=1):
                key = str(value).strip() if value is not None else ""
                if key:
                    header_map[key] = index

            # 若不存在有效表头，则按首条 dict 的 key 建立表头
            if not header_map:
                for index, key in enumerate(dict_rows[0].keys(), start=1):
                    key_text = str(key).strip()
                    if not key_text:
                        continue
                    ws.cell(row=1, column=index).value = key_text
                    header_map[key_text] = index

            # 若新数据有新字段，自动扩展表头到末列
            for row_data in dict_rows:
                for key in row_data.keys():
                    key_text = str(key).strip()
                    if not key_text or key_text in header_map:
                        continue
                    next_col = max(header_map.values(), default=0) + 1
                    ws.cell(row=1, column=next_col).value = key_text
                    header_map[key_text] = next_col

            max_col = max(header_map.values(), default=0)
            for row_data in dict_rows:
                out_row = [None] * max_col
                for key, value in row_data.items():
                    key_text = str(key).strip()
                    col = header_map.get(key_text)
                    if not col:
                        continue
                    out_row[col - 1] = value
                ws.append(out_row)
                append_dict_count += 1

    if set_count == 0 and append_count == 0 and append_dict_count == 0:
        raise ValueError("xlsx_edit 未检测到有效变更（set_cells/append_rows/append_dict_rows 为空）。")

    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.active = workbook.sheetnames.index(ws.title)
    workbook.save(target)
    summary = (
        f"{relative_path}（工作表: {ws.title}，写入单元格 {set_count} 项，"
        f"追加数组行 {append_count} 行，追加字典行 {append_dict_count} 行）"
    )
    if recovery_note:
        summary += f"，{recovery_note}"
    return summary


def _extract_json_block(text: str) -> Optional[Dict[str, Any]]:
    candidates = re.findall(r"\{[\s\S]*\}", text)
    for block in reversed(candidates):
        try:
            parsed = json.loads(block)
            if isinstance(parsed, dict):
                return parsed
        except Exception:
            continue
    return None


def _extract_workspace_plan(text: str) -> Optional[Dict[str, Any]]:
    raw = str(text or "").strip()
    if not raw:
        return None

    parsed = _extract_json_block(raw)
    if isinstance(parsed, dict):
        return parsed

    fence_match = re.search(r"```(?:json)?\s*([\s\S]*?)```", raw, flags=re.IGNORECASE)
    if fence_match:
        fenced = str(fence_match.group(1) or "").strip()
        parsed = _extract_json_block(fenced)
        if isinstance(parsed, dict):
            return parsed
        raw = fenced

    reply_match = re.search(r"[\"']reply[\"']\s*:\s*\"([\s\S]*?)\"", raw, flags=re.IGNORECASE)
    if not reply_match:
        return None

    reply_raw = str(reply_match.group(1) or "")
    reply_text = (
        reply_raw.replace(r"\r\n", "\n")
        .replace(r"\n", "\n")
        .replace(r"\t", "\t")
        .replace(r"\"", '"')
        .replace(r"\\/", "/")
        .replace(r"\\", "\\")
        .strip()
    )
    if not reply_text:
        return None

    return {"reply": reply_text, "actions": []}


def _workspace_execute_actions(root: Path, actions: List[Dict[str, Any]]) -> List[str]:
    logs: List[str] = []
    for item in actions[:12]:
        action = str(item.get("action", "")).strip()
        rel = str(item.get("path", "")).strip()
        try:
            if action == "list_files":
                logs.append("已生成目录清单。")
                continue
            if action == "read_file":
                content = _workspace_read(root, rel)
                preview = content[:600]
                logs.append(f"已读取文件：{rel}\n{preview}")
                continue
            if action == "write_file":
                content = str(item.get("content", ""))
                _workspace_write(root, rel, content)
                logs.append(f"已写入文件：{rel}（{len(content)} 字符）")
                continue
            if action == "append_file":
                content = str(item.get("content", ""))
                _workspace_append(root, rel, content)
                logs.append(f"已追加文件：{rel}（+{len(content)} 字符）")
                continue
            if action == "replace_text":
                old = str(item.get("old", ""))
                new = str(item.get("new", ""))
                replaced = _workspace_replace(root, rel, old, new)
                logs.append(f"已替换文件：{rel}（替换 {replaced} 处）")
                continue
            if action == "xlsx_edit":
                sheet = str(item.get("sheet", "")).strip() or None
                set_cells = item.get("set_cells", [])
                append_rows = item.get("append_rows", [])
                append_dict_rows = item.get("append_dict_rows", [])
                summary = _workspace_xlsx_edit(
                    root,
                    rel,
                    sheet=sheet,
                    set_cells=set_cells,
                    append_rows=append_rows,
                    append_dict_rows=append_dict_rows,
                )
                logs.append(f"已更新 Excel：{summary}")
                continue
            if action == "organize_reimbursement_package":
                package_name = str(item.get("package_name", "")).strip() or None
                options = item.get("options", {}) if isinstance(item.get("options", {}), dict) else {}
                summary = _workspace_prepare_reimbursement_package(root, package_name=package_name, options=options)
                logs.append(summary)
                continue
            logs.append(f"已跳过未知操作：{action}")
        except PermissionError:
            logs.append(
                f"操作失败：{rel} 权限不足或文件被占用。请关闭占用该文件的程序（如 Excel/WPS）后重试。"
            )
        except ValueError as exc:
            logs.append(f"操作失败：{exc}")
        except Exception as exc:
            logs.append(f"操作失败：{rel}，原因：{exc}")
    return logs


def _workspace_result_text(base_reply: str, logs: List[str]) -> str:
    reply = str(base_reply or "").strip() or "已处理你的目录请求。"
    if not logs:
        return reply
    fail_count = sum(1 for item in logs if item.startswith("操作失败"))
    success_count = max(len(logs) - fail_count, 0)
    if fail_count == 0:
        header = f"已完成本次操作（成功 {success_count} 项）。"
    elif success_count == 0:
        header = f"本次操作未成功（失败 {fail_count} 项）。"
    else:
        header = f"已部分完成（成功 {success_count} 项，失败 {fail_count} 项）。"
    return f"{reply}\n\n{header}\n" + "\n".join(f"- {item}" for item in logs)


def _parse_workspace_command(message: str) -> Optional[Dict[str, Any]]:
    text = message.strip()
    # 直连解析：往某个 xlsx 中追加 N 条“学号/姓名”测试数据，避免依赖 LLM 规划不稳定
    xlsx_path_match = re.search(r"[\"“]?([^\"”\n\r]*?\.xlsx)[\"”]?", text, flags=re.IGNORECASE)
    if xlsx_path_match and ("测试数据" in text or "追加" in text):
        target_path = xlsx_path_match.group(1).strip()
        count_match = re.search(r"(\d+)\s*条", text)
        count = int(count_match.group(1)) if count_match else 10
        count = max(1, min(count, 2000))
        sheet_match = re.search(r"\b(Sheet\d+)\b", text, flags=re.IGNORECASE)
        requested_sheet = sheet_match.group(1) if sheet_match else "Sheet1"
        rows = [{"学号": 1000 + i + 1, "姓名": f"测试{i + 1}"} for i in range(count)]
        return {
            "reply": f"已准备向 {target_path} 的 {requested_sheet} 追加 {count} 条测试数据。",
            "actions": [
                {
                    "action": "xlsx_edit",
                    "path": target_path,
                    "sheet": requested_sheet,
                    "append_dict_rows": rows,
                }
            ],
        }

    if text.startswith("/list"):
        return {"reply": "目录如下：", "actions": [{"action": "list_files"}]}

    if text.startswith("/read "):
        rel = text[6:].strip()
        return {"reply": f"读取文件: {rel}", "actions": [{"action": "read_file", "path": rel}]}

    if text.startswith("/write "):
        lines = text.splitlines()
        rel = lines[0][7:].strip()
        content = "\n".join(lines[1:])
        return {
            "reply": f"写入文件: {rel}",
            "actions": [{"action": "write_file", "path": rel, "content": content}],
        }

    if text.startswith("/append "):
        lines = text.splitlines()
        rel = lines[0][8:].strip()
        content = "\n".join(lines[1:])
        return {
            "reply": f"追加文件: {rel}",
            "actions": [{"action": "append_file", "path": rel, "content": content}],
        }

    if text.startswith("/replace "):
        lines = text.splitlines()
        rel = lines[0][9:].strip()
        body = "\n".join(lines[1:])
        marker_old = "---OLD---"
        marker_new = "---NEW---"
        if marker_old in body and marker_new in body:
            old_part = body.split(marker_old, 1)[1]
            old_text, new_text = old_part.split(marker_new, 1)
            return {
                "reply": f"替换文件: {rel}",
                "actions": [
                    {
                        "action": "replace_text",
                        "path": rel,
                        "old": old_text.strip("\n"),
                        "new": new_text.strip("\n"),
                    }
                ],
            }
    package_name_match = re.search(r"([A-Za-z0-9_\-\u4e00-\u9fa5]+\.zip)\b", text, flags=re.IGNORECASE)
    needs_package = (
        "报销" in text
        and any(word in text for word in ["打包", "压缩", "压缩包"])
        and any(word in text for word in ["整理", "材料", "附件", "自动"])
    )
    if needs_package:
        package_name = package_name_match.group(1) if package_name_match else ""
        return {
            "reply": "已开始整理报销材料。若材料齐全会自动生成压缩包；若缺失会先提示你补齐。",
            "actions": [
                {
                    "action": "organize_reimbursement_package",
                    "package_name": package_name,
                }
            ],
        }
    return None


def _run_workspace_agent(
    message: str,
    payload: Dict[str, Any],
    history: List[Dict[str, str]],
    memory_context: str = "",
    referenced_context: str = "",
) -> Dict[str, Any]:
    workspace_root = _safe_workspace_root(payload)
    if workspace_root is None:
        return {
            "ok": False,
            "error": "未绑定有效目录，请先拖拽文件夹到桌宠后再对话。",
        }

    directory_tree = _workspace_tree_text(workspace_root)
    effective_message = _resolve_message_with_referenced_file(message, payload)
    referenced_files = _referenced_file_paths(payload)
    resolved_target_hint = referenced_files[0] if len(referenced_files) == 1 else ""

    command_plan: Optional[Dict[str, Any]] = None

    workspace_task = str(payload.get("workspace_task", "")).strip().lower()
    if workspace_task == "reimbursement_package":
        package_name = str(payload.get("package_name", "")).strip()
        options = payload.get("reimbursement_package_options", {})
        command_plan = {
            "reply": "已按结构化任务开始整理报销材料。若缺失会先提示补齐。",
            "actions": [
                {
                    "action": "organize_reimbursement_package",
                    "package_name": package_name,
                    "options": options if isinstance(options, dict) else {},
                }
            ],
        }

    if command_plan is None:
        # Deterministic fast path: when user selected exactly one file with '@',
        # execute data-append style requests directly to avoid unnecessary path clarification.
        command_plan = _build_direct_plan_from_single_reference(message, payload)

    if command_plan is None:
        command_plan = _parse_workspace_command(effective_message)
    if command_plan is not None:
        logs = _workspace_execute_actions(workspace_root, command_plan.get("actions", []))
        if command_plan.get("actions") and command_plan["actions"][0].get("action") == "list_files":
            return {
                "ok": True,
                "reply": _workspace_result_text(
                    str(command_plan.get("reply", "目录如下：")),
                    [directory_tree],
                ),
                "mode": "workspace",
            }
        return {
            "ok": True,
            "reply": _workspace_result_text(str(command_plan.get("reply", "已处理。")), logs),
            "mode": "workspace",
        }

    if _is_llm_enabled():
        recent_history = history[-8:] if history else []
        history_text = "\n".join(
            f"{item.get('role', 'user')}: {item.get('content', '')}" for item in recent_history
        )
        planner_prompt = textwrap.dedent(
            f"""
            你是本地代码编辑代理，需要在指定目录内操作文件，并向用户提供清晰说明。
            你的首要目标是：准确理解“要改什么、改哪个文件、如何改”，不确定时先澄清，避免误改。
            目录根路径: {workspace_root}
            当前目录树（节选）:
            {directory_tree}

            对话历史（最近）:
            {history_text or '(无)'}

            记忆上下文（可用于保持连续性）:
            {memory_context or '(无)'}

            用户当前引用文件（优先参考）:
            {referenced_context or '(无)'}

            已解析唯一引用目标:
            {resolved_target_hint or '(无)'}

            用户请求:
            {effective_message}

            先在心里完成以下判断，再输出 JSON：
            - 意图类型：问答 / 读文件 / 改文件 / 列目录 / 其他。
            - 目标文件：用户是否明确给了路径、文件名、后缀、模块名、函数名或“这个文件/当前文件”等指代。
            - 改动类型：新增、替换、追加、结构化表格写入。
            - 信息是否充分：如果无法唯一定位文件或改动内容不完整，必须先澄清，不执行写操作。

            文件定位规则（严格执行）：
            1) 用户给了明确相对路径：直接使用该 path。
            2) 用户只给文件名（如 main.ts）：
               - 若目录树中唯一命中，可直接操作；
               - 若有多个同名文件，禁止猜测，reply 里列出候选路径并请用户确认，actions 置空。
            3) 若“已解析唯一引用目标”不为空，且用户说“这个文件/当前文件”，直接将该目标作为 path。
            4) 用户说“这个文件/当前文件”但没有可确定路径时，禁止猜测，reply 里明确要求用户提供相对路径，actions 置空。
            5) 仅当文件可唯一确定时，才允许输出写入类 actions。
            6) 若“已解析唯一引用目标”不为空，不要再向用户追问路径，直接执行或仅追问业务字段。

            请仅返回 JSON 对象，不要加解释文字。格式：
            {{
              "reply": "给用户的简短说明（先说理解到的目标；若可执行，说明将做什么；若信息不足，明确要补充什么）",
              "actions": [
                {{"action": "list_files"}},
                {{"action": "read_file", "path": "relative/path"}},
                {{"action": "write_file", "path": "relative/path", "content": "..."}},
                {{"action": "append_file", "path": "relative/path", "content": "..."}},
                {{"action": "replace_text", "path": "relative/path", "old": "...", "new": "..."}},
                {{
                  "action": "xlsx_edit",
                  "path": "relative/path.xlsx",
                  "sheet": "Sheet1",
                  "set_cells": [{{"cell": "A1", "value": "标题"}}],
                  "append_rows": [[1, "张三"], [2, "李四"]],
                  "append_dict_rows": [
                    {{"学号": 101, "姓名": "张三"}},
                    {{"学号": 202, "姓名": "李四"}}
                  ]
                }}
              ]
            }}

            规则：
            1) path 必须是相对路径。
            2) 若需要改文件，优先 replace_text；若文件不存在再 write_file。
            3) 若用户只是询问，则 actions 可为空。
            4) 若信息不足（目标文件不唯一/改动描述不完整），必须先提问澄清，actions 必须为空。
            5) reply 使用用户可读语言，不要只输出技术术语；若执行失败，需要说明原因和可替代方案。
            6) 若执行改动，reply 需要简要说明改动范围（改了哪些文件/内容）。
            7) .xlsx/.xlsm 文件只能用 xlsx_edit，严禁用 write_file、append_file、replace_text 进行文本写入。
            8) .xls/.docx/.pdf/图片/压缩包等二进制文件，禁止文本写入；若用户要求编辑，先说明限制并给可行替代方案。
            9) 若用户给的是“按字段”的结构化数据，优先使用 append_dict_rows 按表头写入。
            """
        ).strip()

        planner_raw = _llm_chat(message=planner_prompt, history=[], kb_context="")
        parsed = _extract_workspace_plan(planner_raw)
        if parsed is None:
            fallback = str(planner_raw or "").strip()
            if fallback.startswith("{") and "reply" in fallback and "actions" in fallback:
                fallback = "我没有拿到可执行计划。请提供要编辑的相对路径和具体修改内容。"
            return {
                "ok": True,
                "reply": fallback or "我没有拿到可执行计划。请重试或补充更明确的修改目标。",
                "mode": "workspace",
            }

        actions = parsed.get("actions", [])
        safe_actions = [item for item in actions if isinstance(item, dict)] if isinstance(actions, list) else []
        logs = _workspace_execute_actions(workspace_root, safe_actions)

        if any(str(item.get("action", "")).strip() == "list_files" for item in safe_actions):
            logs.append(directory_tree)

        reply = str(parsed.get("reply", "已完成目录操作。"))
        reply = _workspace_result_text(reply, logs)
        return {
            "ok": True,
            "reply": reply,
            "mode": "workspace",
        }

    return {
        "ok": True,
        "mode": "workspace",
        "reply": (
            "当前未启用 LLM 规划。可用命令：\n"
            "/list\n"
            "/read 相对路径\n"
            "/write 相对路径 + 换行后文件内容\n"
            "/append 相对路径 + 换行后追加内容\n"
            "/replace 相对路径 + 换行后使用 ---OLD--- 与 ---NEW--- 标记"
        ),
    }


def _extract_task_request(message: str, payload: Dict[str, Any]) -> Tuple[Optional[str], Dict[str, Any]]:
    alias_map = {
        "t1_qa": "qa",
        "t2_recon": "recon",
        "t3_material": "reimburse",
        "t4_budget_fill": "budget",
        "t5_final_fill": "final_account",
        "t6_file_edit": "file_edit",
    }
    task_type = str(payload.get("task_type", "")).strip().lower()
    task_type = alias_map.get(task_type, task_type)
    task_payload = payload.get("task_payload", payload)
    if isinstance(task_payload, dict) and task_type:
        return task_type, task_payload

    if message.startswith("/task "):
        parts = message.split(maxsplit=2)
        task_type = parts[1].strip().lower() if len(parts) > 1 else ""
        task_type = alias_map.get(task_type, task_type)
        if task_type:
            return task_type, payload

    return None, payload


def _supervisor_infer_task(message: str, payload: Dict[str, Any]) -> Tuple[Optional[str], float, List[str]]:
    text = str(message or "").strip().lower()
    reasons: List[str] = []
    if not text:
        return None, 0.0, reasons

    if any(key in text for key in ("修改文件", "编辑文件", "替换文本", "write_file", "replace_text", "xlsx_edit")):
        reasons.append("R602_FILE_EDIT")
        return "file_edit", 0.93, reasons

    has_budget = any(key in text for key in ("预算", "budget"))
    has_final = any(key in text for key in ("决算", "final", "结项"))
    has_check = any(key in text for key in ("核对", "比对", "差异", "一致性"))
    has_fill = any(key in text for key in ("填写", "回填", "填报"))

    if has_budget and has_final and has_check:
        reasons.append("R201_RECON")
        return "recon", 0.9, reasons
    if has_budget and has_fill:
        reasons.append("R401_BUDGET_FILL")
        return "budget", 0.88, reasons
    if has_final and has_fill:
        reasons.append("R501_FINAL_FILL")
        return "final_account", 0.88, reasons
    if any(key in text for key in ("整理材料", "报销材料", "附件整理", "打包", "归档")):
        reasons.append("R301_MATERIAL")
        return "reimburse", 0.86, reasons
    if any(key in text for key in ("报销规则", "能不能报", "附件要求", "制度", "口径")):
        reasons.append("R101_QA")
        return "qa", 0.85, reasons
    if has_budget:
        reasons.append("R402_BUDGET")
        return "budget", 0.73, reasons
    if has_final:
        reasons.append("R502_FINAL")
        return "final_account", 0.73, reasons
    if any(key in text for key in ("报销", "发票", "附件")):
        reasons.append("R302_REIMBURSE")
        return "qa", 0.72, reasons
    if bool(payload.get("workspace_mode", False)):
        reasons.append("R603_WORKSPACE")
        return "file_edit", 0.66, reasons
    return None, 0.0, reasons


def _looks_like_workspace_intent(message: str) -> bool:
    text = str(message or "").strip()
    if not text:
        return False

    lowered = text.lower()
    if text.startswith(("/list", "/read ", "/write ", "/append ", "/replace ")):
        return True

    if _parse_workspace_command(text) is not None:
        return True

    path_or_file_pattern = (
        r"([A-Za-z]:\\|/|\.{1,2}[\\/])?[^\s\"'<>|?:*]+"
        r"\.(py|ts|tsx|js|jsx|json|md|txt|yaml|yml|toml|ini|csv|xlsx|xlsm|docx|pdf)\b"
    )
    if re.search(path_or_file_pattern, text, flags=re.IGNORECASE):
        return True

    file_keywords = (
        "文件",
        "目录",
        "路径",
        "代码",
        "仓库",
        "workspace",
        "readme",
        ".gitignore",
        "main.py",
        "main.ts",
    )
    action_keywords = (
        "读取",
        "查看",
        "打开",
        "编辑",
        "修改",
        "替换",
        "追加",
        "写入",
        "新建",
        "创建",
        "删除",
        "重命名",
        "列出",
        "搜索",
        "查找",
    )
    has_file_hint = any(word in lowered for word in file_keywords)
    has_action_hint = any(word in text for word in action_keywords) or any(word in lowered for word in action_keywords)
    return bool(has_file_hint and has_action_hint)


def _route_request_mode(message: str, payload: Dict[str, Any]) -> Tuple[str, Optional[str], Dict[str, Any]]:
    override_mode = str(payload.get("route_mode", "")).strip().lower()
    task_type, task_payload = _extract_task_request(message, payload)

    if override_mode == "task":
        return "task", task_type, task_payload
    if override_mode == "workspace":
        enriched_payload = {**payload, "workspace_mode": True}
        return "task", "auto", enriched_payload
    if override_mode == "chat":
        return "task", "auto", payload

    if task_type:
        return "task", task_type, task_payload

    # Unify routing: implicit requests always enter the graph as auto tasks,
    # then IntentNode becomes the only task-type decision source.
    return "task", "auto", payload


def _prepare_task_payload_for_dispatch(
    message: str,
    base_payload: Dict[str, Any],
    task_payload: Dict[str, Any],
    task_type: str,
) -> Dict[str, Any]:
    merged_payload: Dict[str, Any] = {}
    if isinstance(base_payload, dict):
        merged_payload.update(base_payload)
    if isinstance(task_payload, dict):
        merged_payload.update(task_payload)

    merged_payload["query"] = str(merged_payload.get("query", "")).strip() or str(message or "").strip()

    workspace_dir = str(
        merged_payload.get("workspace_dir", "") or merged_payload.get("workspace_root", "")
    ).strip()
    if workspace_dir:
        merged_payload["workspace_dir"] = workspace_dir
        merged_payload.setdefault("workspace_root", workspace_dir)

    referenced_files = _referenced_file_paths(merged_payload)
    should_prepare_file_actions = (
        task_type == "file_edit"
        or bool(merged_payload.get("workspace_mode", False))
        or len(referenced_files) > 0
        or _looks_like_workspace_intent(message)
    )
    if should_prepare_file_actions:
        effective_message = _resolve_message_with_referenced_file(message, merged_payload)
        if not isinstance(merged_payload.get("actions"), list) or not merged_payload.get("actions"):
            command_plan = _build_direct_plan_from_single_reference(message, merged_payload)
            if command_plan is None:
                command_plan = _parse_workspace_command(effective_message)
            if isinstance(command_plan, dict):
                actions = command_plan.get("actions", [])
                if isinstance(actions, list):
                    merged_payload["actions"] = [item for item in actions if isinstance(item, dict)]

        policy = merged_payload.get("policy", {})
        if not isinstance(policy, dict):
            policy = {}
        if "requires_confirmation" not in policy:
            policy["requires_confirmation"] = True
        merged_payload["policy"] = policy

    return merged_payload


def _get_llm_base_url() -> str:
    raw = (
        os.getenv("AGENT_LLM_BASE_URL", "").strip()
        or os.getenv("AGENT_LLM_API_URL", "").strip()
        or "https://api.openai.com/v1"
    )
    normalized = raw.rstrip("/")
    parsed = urlparse(normalized)

    path = (parsed.path or "").rstrip("/")
    if not path:
        normalized = f"{normalized}/v1"

    return normalized


def _safe_int_env(name: str, default: int, *, min_value: int) -> int:
    raw = os.getenv(name, str(default)).strip()
    try:
        value = int(raw)
    except ValueError:
        return default
    return max(value, min_value)


def _normalize_history(history: List[Dict[str, str]], message: str) -> List[Dict[str, str]]:
    history_messages: List[Dict[str, str]] = []
    for item in history[-20:]:
        role = str(item.get("role", "")).strip()
        content = str(item.get("content", "")).strip()
        if role not in {"user", "assistant"} or not content:
            continue
        history_messages.append({"role": role, "content": content})

    normalized: List[Dict[str, str]] = []
    for item in history_messages:
        if not normalized:
            if item["role"] != "user":
                continue
            normalized.append(item)
            continue

        if normalized[-1]["role"] == item["role"]:
            normalized[-1]["content"] += "\n\n" + item["content"]
        else:
            normalized.append(item)

    if not normalized:
        return [{"role": "user", "content": message}]

    if normalized[-1]["role"] != "user":
        normalized.append({"role": "user", "content": message})
    elif normalized[-1]["content"] != message:
        normalized[-1]["content"] += "\n\n" + message
    return normalized


def _build_llm_messages(message: str, history: List[Dict[str, str]], kb_context: str) -> Tuple[List[Dict[str, str]], bool]:
    base_url = _get_llm_base_url()
    parsed = urlparse(base_url)
    host = (parsed.hostname or "").lower()
    is_local = host in {"localhost", "127.0.0.1", "::1"}

    normalized = _normalize_history(history, message)

    messages: List[Dict[str, str]] = []
    if is_local:
        if normalized:
            context_block = f"\n\n可参考的知识库片段（优先基于这些资料回答）：\n{kb_context}" if kb_context else ""
            normalized[0]["content"] = f"{SYSTEM_PROMPT}{context_block}\n\n用户问题：{normalized[0]['content']}"
        messages.extend(normalized)
    else:
        system_prompt = SYSTEM_PROMPT
        if kb_context:
            system_prompt += f"\n\n可参考的知识库片段：\n{kb_context}"
        messages.append({"role": "system", "content": system_prompt})
        messages.extend(normalized)

    return messages, is_local


def _rule_reply(message: str, payload: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    budget_source = payload.get("budget_source")
    actual_source = payload.get("actual_source")

    if budget_source is not None and actual_source is not None:
        return _run_audit(budget_source, actual_source)

    if re.search(r"sample|示例|demo", message, flags=re.IGNORECASE):
        from agent.sample_data import get_sample_payloads  # noqa: WPS433

        budget_json, actual_json = get_sample_payloads()
        return _run_audit(budget_json, actual_json)

    if "高风险" in message or "风险" in message:
        return {
            "reply": "高风险触发规则：类目无法映射、单项超支>10%、总额超预算、餐饮/会议缺签到或通知附件。",
        }

    if "材料" in message or "附件" in message:
        return {
            "reply": "餐饮/会议类支出需具备签到表或通知文件提示，建议在上传时同时附发票和明细。",
        }

    return None


def _brief_report(report_json: Dict[str, Any]) -> str:
    summary = report_json.get("summary", {})
    total = summary.get("total_issues", 0)
    high = summary.get("high_risk_issues", 0)
    status = summary.get("overall_status", "UNKNOWN")
    return f"审计完成：状态={status}，问题总数={total}，高风险={high}。"


def _run_audit(budget_source: Any, actual_source: Any) -> Dict[str, Any]:
    try:
        from agent.graph_builder import build_graph  # noqa: WPS433
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "审计模式依赖缺失，请在当前 Python 环境安装 requirements.txt（含 pandas/langgraph/jsonschema）。"
        ) from exc

    app = build_graph()
    state: Dict[str, Any] = {
        "budget_source": budget_source,
        "actual_source": actual_source,
        "discrepancies": [],
        "suggestions": [],
    }
    result = app.invoke(state)
    report = result.get("report", {})
    report_json = report.get("report_json", {})
    report_markdown = report.get("report_markdown", "")
    return {
        "reply": _brief_report(report_json),
        "report_json": report_json,
        "report_markdown": report_markdown,
    }


def _format_task_reply(task_type: str, result: Dict[str, Any]) -> str:
    normalized_task = str(task_type or "").strip().lower()
    result_type = str(result.get("type", "")).strip().lower()

    if normalized_task == "recon" or result_type == "recon":
        summary = result.get("summary", {}) if isinstance(result.get("summary", {}), dict) else {}
        total_items = int(summary.get("total_items", 0) or 0)
        blocking = int(summary.get("blocking", 0) or 0)
        warning = int(summary.get("warning", 0) or 0)
        hint = int(summary.get("hint", 0) or 0)
        status = str(result.get("status", "unknown"))
        if status == "needs_clarification":
            return str(
                result.get(
                    "message",
                    "核对数据不足，请补充预算与决算数据后重试。",
                )
            )
        base = (
            f"预算/决算核对完成：状态={status}，共核对 {total_items} 项，"
            f"阻断 {blocking} 项，预警 {warning} 项，提示 {hint} 项。"
        )
        raw_limit = result.get("detail_limit", 3)
        try:
            detail_limit = max(1, min(int(raw_limit), 10))
        except (TypeError, ValueError):
            detail_limit = 3

        def _to_lines(items: Any, title: str, limit: int) -> List[str]:
            rows = items if isinstance(items, list) else []
            lines: List[str] = []
            for item in rows[:limit]:
                if not isinstance(item, dict):
                    continue
                key = str(item.get("key", "")).strip() or "未命名项"
                abs_diff = item.get("abs_diff", 0)
                pct_diff = item.get("pct_diff", 0)
                reason = str(item.get("reason", "")).strip() or "无"
                lines.append(f"- {key}：差额={abs_diff}，差异率={pct_diff}，原因={reason}")
            if not lines:
                return []
            return [title] + lines

        def _build_suggestions(blocking_items: Any, warning_items: Any) -> List[str]:
            suggestions: List[str] = []
            blocking_rows = blocking_items if isinstance(blocking_items, list) else []
            warning_rows = warning_items if isinstance(warning_items, list) else []
            joined_reasons = " ".join(
                str(item.get("reason", "")) for item in blocking_rows + warning_rows if isinstance(item, dict)
            )
            custom_rules = result.get("suggestion_rules", [])
            if isinstance(custom_rules, list):
                for rule in custom_rules[:12]:
                    if not isinstance(rule, dict):
                        continue
                    reason_contains = rule.get("reason_contains", [])
                    if isinstance(reason_contains, str):
                        reason_tokens = [reason_contains]
                    elif isinstance(reason_contains, list):
                        reason_tokens = [str(item).strip() for item in reason_contains if str(item).strip()]
                    else:
                        reason_tokens = []
                    suggestion = str(rule.get("suggestion", "")).strip()
                    if not suggestion:
                        continue
                    if reason_tokens and not any(token in joined_reasons for token in reason_tokens):
                        continue
                    if suggestion not in suggestions:
                        suggestions.append(suggestion)

            if any("缺少对应项" in str(item.get("reason", "")) for item in blocking_rows if isinstance(item, dict)):
                suggestions.append("补齐预算或决算缺失项后重新核对，确保主键维度（期间/部门/科目/项目）一致。")
            if any("阻断阈值" in str(item.get("reason", "")) for item in blocking_rows if isinstance(item, dict)):
                suggestions.append("针对阻断项优先复核金额来源与汇总口径，必要时暂停自动回填。")
            if any("预警阈值" in str(item.get("reason", "")) for item in warning_rows if isinstance(item, dict)):
                suggestions.append("预警项建议提供差异说明，并在审批备注中记录原因。")
            if "口径" in joined_reasons or "币种" in joined_reasons or "税率" in joined_reasons:
                suggestions.append("检查币种、税率、含税口径是否一致。")
            if not suggestions and (blocking_rows or warning_rows):
                suggestions.append("按差异明细逐项复核原始凭证和表内公式，确认后再提交。")
            return suggestions[:3]

        detail_lines: List[str] = []
        blocking_items = result.get("blocking_items", [])
        warning_items = result.get("warning_items", [])
        detail_lines.extend(_to_lines(blocking_items, "阻断项明细：", detail_limit))
        detail_lines.extend(_to_lines(warning_items, "预警项明细：", detail_limit))
        suggestions = _build_suggestions(blocking_items, warning_items)
        if suggestions:
            detail_lines.append("建议处理：")
            detail_lines.extend(f"- {item}" for item in suggestions)
        if detail_lines:
            return base + "\n" + "\n".join(detail_lines)
        return base

    if normalized_task == "file_edit" or result_type == "file_edit":
        status = str(result.get("status", "unknown"))
        changeset = result.get("changeset", [])
        change_count = len(changeset) if isinstance(changeset, list) else 0
        return f"文件编辑任务完成：状态={status}，变更 {change_count} 项。"

    if normalized_task == "qa" or result_type == "qa":
        answer = str(result.get("answer", "")).strip()
        return answer or "问答任务已完成。"

    return f"任务已完成：{task_type}"


def _run_v2_task(task_type: str, task_payload: Dict[str, Any]) -> Dict[str, Any]:
    try:
        from agent import EventBus, TaskDispatcher  # noqa: WPS433
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "任务调度模式依赖缺失，请确认已安装项目依赖并同步到当前 Python 环境。"
        ) from exc

    event_bus = EventBus()
    progress_events: List[Dict[str, Any]] = []
    event_bus.subscribe("task_progress", lambda evt: progress_events.append(evt))
    dispatcher = TaskDispatcher(event_bus)
    result = dispatcher.dispatch(task_type, task_payload)
    reply = _format_task_reply(task_type, result if isinstance(result, dict) else {})
    return {
        "reply": reply,
        "mode": "task",
        "task_type": task_type,
        "task_result": result,
        "task_progress": progress_events,
    }


def _help_text() -> str:
    return (
        "你可以这样和我对话：\n"
        "1) 输入“运行sample审计”触发内置示例审计；\n"
        "2) 输入“如何修复高风险问题”等规则咨询；\n"
        "3) 传入 payload.budget_source / payload.actual_source 做真实数据审计；\n"
        "4) 传入 payload.task_type（qa/reimburse/final_account/budget/recon/file_edit/sandbox_exec）触发新图任务。"
        "\n5) 传入 payload.workspace_mode=true + payload.workspace_dir 使用目录编辑工具模式。"
    )


def _is_llm_enabled() -> bool:
    base_url = _get_llm_base_url()
    api_key = os.getenv("AGENT_LLM_API_KEY", "").strip()
    parsed = urlparse(base_url)
    host = (parsed.hostname or "").lower()
    is_local = host in {"localhost", "127.0.0.1", "::1"}
    return bool(api_key) or is_local


def _get_kb_context(message: str) -> str:
    kb_path = Path(os.getenv("AGENT_KB_PATH", str(DEFAULT_KB_PATH))).resolve()
    top_k = _safe_int_env("AGENT_KB_TOP_K", 4, min_value=1)
    max_chars = _safe_int_env("AGENT_KB_MAX_CHARS", 1800, min_value=600)

    if not kb_path.exists():
        return ""

    try:
        from agent.kb.retriever import format_retrieved_context, retrieve_chunks, search_policy  # noqa: WPS433
    except ModuleNotFoundError:
        return ""

    try:
        # Prefer hybrid semantic retrieval; fallback to keyword retrieval for robustness.
        chunks = search_policy(message, top_k=top_k, kb_path=kb_path)
        if not chunks:
            chunks = retrieve_chunks(message, kb_path=kb_path, top_k=top_k)
        return format_retrieved_context(chunks, max_chars=max_chars)
    except Exception:
        return ""


def _llm_chat(message: str, history: List[Dict[str, str]], kb_context: str = "") -> str:
    base_url = _get_llm_base_url()
    api_key = os.getenv("AGENT_LLM_API_KEY", "").strip()
    parsed = urlparse(base_url)
    host = (parsed.hostname or "").lower()
    is_local = host in {"localhost", "127.0.0.1", "::1"}

    if not api_key and not is_local:
        raise ValueError("未配置 AGENT_LLM_API_KEY。非本地 LLM 服务需要有效 API Key。")

    model = os.getenv("AGENT_LLM_MODEL", "gpt-4o-mini").strip()
    timeout_seconds = _safe_int_env("AGENT_LLM_TIMEOUT", 60, min_value=10)
    messages, _ = _build_llm_messages(message=message, history=history, kb_context=kb_context)

    body = {
        "model": model,
        "messages": messages,
        "temperature": 0.2,
    }

    headers = {
        "Content-Type": "application/json",
    }
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"

    req = urllib.request.Request(
        url=f"{base_url}/chat/completions",
        data=json.dumps(body).encode("utf-8"),
        headers=headers,
        method="POST",
    )

    try:
        with urllib.request.urlopen(req, timeout=timeout_seconds) as resp:
            raw = resp.read().decode("utf-8")
            payload = json.loads(raw)
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore") if exc.fp else str(exc)
        raise RuntimeError(f"LLM 接口请求失败: HTTP {exc.code} - {detail}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"LLM 接口网络错误: {exc}") from exc

    choices = payload.get("choices", [])
    if not choices:
        raise RuntimeError("LLM 返回为空，未获取到回答。")

    content = choices[0].get("message", {}).get("content", "")
    text = str(content).strip()
    if not text:
        raise RuntimeError("LLM 返回内容为空。")
    return text


def _llm_chat_stream(message: str, history: List[Dict[str, str]], kb_context: str = "") -> Iterator[str]:
    base_url = _get_llm_base_url()
    api_key = os.getenv("AGENT_LLM_API_KEY", "").strip()
    parsed = urlparse(base_url)
    host = (parsed.hostname or "").lower()
    is_local = host in {"localhost", "127.0.0.1", "::1"}

    if not api_key and not is_local:
        raise ValueError("未配置 AGENT_LLM_API_KEY。非本地 LLM 服务需要有效 API Key。")

    model = os.getenv("AGENT_LLM_MODEL", "gpt-4o-mini").strip()
    timeout_seconds = _safe_int_env("AGENT_LLM_TIMEOUT", 60, min_value=10)
    messages, _ = _build_llm_messages(message=message, history=history, kb_context=kb_context)

    body = {
        "model": model,
        "messages": messages,
        "temperature": 0.2,
        "stream": True,
    }

    headers = {
        "Content-Type": "application/json",
    }
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"

    req = urllib.request.Request(
        url=f"{base_url}/chat/completions",
        data=json.dumps(body).encode("utf-8"),
        headers=headers,
        method="POST",
    )

    try:
        with urllib.request.urlopen(req, timeout=timeout_seconds) as resp:
            for raw_line in resp:
                line = raw_line.decode("utf-8", errors="ignore").strip()
                if not line or not line.startswith("data:"):
                    continue
                data = line[5:].strip()
                if data == "[DONE]":
                    break

                try:
                    payload = json.loads(data)
                except json.JSONDecodeError:
                    continue

                choices = payload.get("choices", [])
                if not choices:
                    continue
                delta = choices[0].get("delta", {}).get("content", "")
                text = str(delta)
                if text:
                    yield text
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore") if exc.fp else str(exc)
        raise RuntimeError(f"LLM 接口请求失败: HTTP {exc.code} - {detail}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"LLM 接口网络错误: {exc}") from exc


def _iter_text_chunks(text: str, chunk_size: int = 36) -> Iterator[str]:
    content = text or ""
    if not content:
        return
    for index in range(0, len(content), max(chunk_size, 8)):
        yield content[index : index + max(chunk_size, 8)]


def handle_request_stream(request: Dict[str, Any]) -> Iterator[Dict[str, Any]]:
    message = str(request.get("message", "")).strip()
    raw_payload = request.get("payload", {}) or {}
    payload = raw_payload if isinstance(raw_payload, dict) else {}
    history = payload.get("history", []) if isinstance(payload, dict) else []
    safe_history = history if isinstance(history, list) else []
    if bool(payload.get("memory_reset", False)):
        _reset_memory_session(payload)
    memory_ctx = _memory_context(payload)
    referenced_ctx = _referenced_files_context(payload)

    route_mode, task_type, task_payload = _route_request_mode(message, payload)

    if route_mode == "task":
        effective_task_type = str(task_type or "").strip() or "auto"
        prepared_task_payload = _prepare_task_payload_for_dispatch(
            message=message,
            base_payload=payload,
            task_payload=task_payload if isinstance(task_payload, dict) else {},
            task_type=effective_task_type,
        )
        yield {"type": "status", "status": f"正在执行任务: {effective_task_type}"}
        task_resp = _run_v2_task(effective_task_type, prepared_task_payload)
        for step in task_resp.get("task_progress", []):
            step_name = str(step.get("step", ""))
            tool_name = str(step.get("tool_name", ""))
            yield {"type": "status", "status": f"步骤: {step_name} | Tool: {tool_name}"}
        reply = str(task_resp.get("reply", "任务完成"))
        _remember_turn(payload, message, reply)
        for chunk in _iter_text_chunks(reply):
            yield {"type": "delta", "delta": chunk}
        yield {"type": "done", "response": {"ok": True, **task_resp}}
        return

    yield {"type": "status", "status": "正在分析意图..."}
    rule_result = _rule_reply(message, payload)
    if rule_result is not None:
        yield {"type": "status", "status": "正在处理审计规则..."}
        reply = str(rule_result.get("reply", ""))
        report_markdown = str(rule_result.get("report_markdown", "") or "")
        _remember_turn(payload, message, reply + (f"\n\n{report_markdown}" if report_markdown else ""))
        for chunk in _iter_text_chunks(reply):
            yield {"type": "delta", "delta": chunk}
        if report_markdown:
            yield {"type": "delta", "delta": f"\n\n{report_markdown}"}
        yield {"type": "done", "response": {"ok": True, **rule_result}}
        return

    if _is_llm_enabled():
        yield {"type": "status", "status": "正在调用 RAG 知识库检索..."}
        kb_context = _merge_context_blocks(_get_kb_context(message), memory_ctx, referenced_ctx)
        yield {"type": "status", "status": "正在生成回答..."}
        streamed_reply = ""
        try:
            for chunk in _llm_chat_stream(message=message, history=safe_history, kb_context=kb_context):
                streamed_reply += chunk
                yield {"type": "delta", "delta": chunk}
        except Exception:
            streamed_reply = _llm_chat(message=message, history=safe_history, kb_context=kb_context)
            for chunk in _iter_text_chunks(streamed_reply):
                yield {"type": "delta", "delta": chunk}

        _remember_turn(payload, message, streamed_reply)
        yield {"type": "done", "response": {"ok": True, "reply": streamed_reply, "mode": "llm"}}
        return

    reply = _help_text()
    _remember_turn(payload, message, reply)
    for chunk in _iter_text_chunks(reply):
        yield {"type": "delta", "delta": chunk}
    yield {"type": "done", "response": {"ok": True, "reply": reply}}


def handle_request(request: Dict[str, Any]) -> Dict[str, Any]:
    message = str(request.get("message", "")).strip()
    raw_payload = request.get("payload", {}) or {}
    payload = raw_payload if isinstance(raw_payload, dict) else {}
    history = payload.get("history", []) if isinstance(payload, dict) else []
    safe_history = history if isinstance(history, list) else []
    if bool(payload.get("memory_reset", False)):
        _reset_memory_session(payload)
    memory_ctx = _memory_context(payload)
    referenced_ctx = _referenced_files_context(payload)

    route_mode, task_type, task_payload = _route_request_mode(message, payload)

    if route_mode == "task":
        effective_task_type = str(task_type or "").strip() or "auto"
        prepared_task_payload = _prepare_task_payload_for_dispatch(
            message=message,
            base_payload=payload,
            task_payload=task_payload if isinstance(task_payload, dict) else {},
            task_type=effective_task_type,
        )
        result = {"ok": True, **_run_v2_task(effective_task_type, prepared_task_payload)}
        _remember_turn(payload, message, str(result.get("reply", "")))
        return result

    rule_result = _rule_reply(message, payload)
    if rule_result is not None:
        result = {"ok": True, **rule_result}
        _remember_turn(payload, message, str(result.get("reply", "")))
        return result

    if _is_llm_enabled():
        kb_context = _merge_context_blocks(_get_kb_context(message), memory_ctx, referenced_ctx)
        llm_reply = _llm_chat(message=message, history=safe_history, kb_context=kb_context)
        _remember_turn(payload, message, llm_reply)
        return {"ok": True, "reply": llm_reply, "mode": "llm"}

    fallback = _help_text()
    _remember_turn(payload, message, fallback)
    return {"ok": True, "reply": fallback}


def _configure_stdio() -> None:
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8")
        except Exception:
            continue


def _safe_write_line(line: str) -> bool:
    try:
        sys.stdout.write(line + "\n")
        sys.stdout.flush()
        return True
    except BrokenPipeError:
        return False
    except OSError as exc:
        if exc.errno in {errno.EPIPE, errno.ECONNRESET}:
            return False
        raise


def _emit_json(payload: Dict[str, Any]) -> bool:
    return _safe_write_line(json.dumps(payload, ensure_ascii=False))


def _handle_request_payload(request: Dict[str, Any]) -> bool:
    if request.get("command") == "shutdown":
        _emit_json({"type": "status", "status": "shutdown"})
        return False

    if bool(request.get("stream", False)):
        for event in handle_request_stream(request):
            if not _emit_json(event):
                return False
        return True

    response = handle_request(request)
    return _emit_json(response)


def _handle_raw_request(raw: str) -> bool:
    if not raw:
        return True
    try:
        request = json.loads(raw)
        if not isinstance(request, dict):
            raise ValueError("request payload must be a JSON object")
        return _handle_request_payload(request)
    except Exception as exc:  # pragma: no cover
        if "request" in locals() and isinstance(request, dict) and bool(request.get("stream", False)):
            return _emit_json({"type": "error", "error": str(exc)})
        return _emit_json({"ok": False, "error": str(exc)})


def main() -> None:
    _configure_stdio()

    shutdown_requested = {"value": False}

    def _request_shutdown(signum: int, frame: Any) -> None:
        shutdown_requested["value"] = True

    for sig in (signal.SIGINT, signal.SIGTERM):
        try:
            signal.signal(sig, _request_shutdown)
        except Exception:
            continue

    for raw_line in sys.stdin:
        if shutdown_requested["value"]:
            break
        raw_line = raw_line.strip()
        if not raw_line:
            continue
        if not _handle_raw_request(raw_line):
            break


if __name__ == "__main__":
    main()
