from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Tuple

from agent.graphs.state import AppState


TEXT_BLOCKED_SUFFIXES = {
    ".xlsx",
    ".xls",
    ".xlsm",
    ".xlsb",
    ".docx",
    ".doc",
    ".pptx",
    ".ppt",
    ".pdf",
    ".png",
    ".jpg",
    ".jpeg",
    ".gif",
    ".webp",
    ".bmp",
    ".zip",
    ".rar",
    ".7z",
}

XLSX_EDIT_SUFFIXES = {".xlsx", ".xlsm"}
HIGH_RISK_ACTIONS = {"write_file", "append_file", "replace_text", "xlsx_edit", "material_package"}


def file_edit_start_node(state: AppState) -> AppState:
    return {"task_progress": state.get("task_progress", []) + [{"step": "file_edit_start", "tool_name": "start"}]}


def _safe_target(root: Path, relative_path: str) -> Path:
    rel = str(relative_path or "").strip().replace("\\", "/")
    if not rel:
        raise ValueError("path 不能为空")
    target = (root / rel).resolve()
    try:
        target.relative_to(root)
    except ValueError as exc:
        raise ValueError("禁止访问 workspace 外路径") from exc
    return target


def _append_changeset(changeset: List[Dict[str, Any]], path: str, kind: str, summary: str) -> None:
    changeset.append({"path": path, "kind": kind, "summary": summary})


def _execute_actions(root: Path, actions: List[Dict[str, Any]]) -> Tuple[List[str], List[Dict[str, Any]], List[str]]:
    logs: List[str] = []
    changeset: List[Dict[str, Any]] = []
    errors: List[str] = []

    for item in actions[:20]:
        action = str(item.get("action", "")).strip()
        rel = str(item.get("path", "")).strip()
        try:
            if action == "read_file":
                target = _safe_target(root, rel)
                if target.suffix.lower() in TEXT_BLOCKED_SUFFIXES:
                    raise ValueError(f"不支持读取该文件类型: {target.name}")
                if not target.exists():
                    raise FileNotFoundError(f"文件不存在: {rel}")
                content = target.read_text(encoding="utf-8")
                preview = content[:200]
                logs.append(f"已读取 {rel}: {preview}")
                continue

            if action == "write_file":
                target = _safe_target(root, rel)
                if target.suffix.lower() in TEXT_BLOCKED_SUFFIXES:
                    raise ValueError(f"不支持文本写入该文件类型: {target.name}")
                content = str(item.get("content", ""))
                target.parent.mkdir(parents=True, exist_ok=True)
                target.write_text(content, encoding="utf-8")
                _append_changeset(changeset, rel, action, f"写入 {len(content)} 字符")
                logs.append(f"已写入 {rel}")
                continue

            if action == "append_file":
                target = _safe_target(root, rel)
                if target.suffix.lower() in TEXT_BLOCKED_SUFFIXES:
                    raise ValueError(f"不支持文本追加该文件类型: {target.name}")
                content = str(item.get("content", ""))
                target.parent.mkdir(parents=True, exist_ok=True)
                with target.open("a", encoding="utf-8") as file_obj:
                    file_obj.write(content)
                _append_changeset(changeset, rel, action, f"追加 {len(content)} 字符")
                logs.append(f"已追加 {rel}")
                continue

            if action == "replace_text":
                target = _safe_target(root, rel)
                if target.suffix.lower() in TEXT_BLOCKED_SUFFIXES:
                    raise ValueError(f"不支持文本替换该文件类型: {target.name}")
                if not target.exists():
                    raise FileNotFoundError(f"文件不存在: {rel}")
                old = str(item.get("old", ""))
                new = str(item.get("new", ""))
                source = target.read_text(encoding="utf-8")
                count = source.count(old) if old else 0
                target.write_text(source.replace(old, new), encoding="utf-8")
                _append_changeset(changeset, rel, action, f"替换 {count} 处")
                logs.append(f"已替换 {rel}")
                continue

            if action == "xlsx_edit":
                target = _safe_target(root, rel)
                if target.suffix.lower() not in XLSX_EDIT_SUFFIXES:
                    raise ValueError("xlsx_edit 仅支持 .xlsx/.xlsm")
                try:
                    import openpyxl  # type: ignore
                except Exception as exc:
                    raise RuntimeError("缺少 openpyxl 依赖") from exc

                if target.exists():
                    workbook = openpyxl.load_workbook(target)
                else:
                    workbook = openpyxl.Workbook()
                sheet_name = str(item.get("sheet", "")).strip()
                ws = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
                if sheet_name and sheet_name not in workbook.sheetnames:
                    ws.title = sheet_name

                set_cells = item.get("set_cells", [])
                append_rows = item.get("append_rows", [])
                append_dict_rows = item.get("append_dict_rows", [])
                set_count = 0
                append_count = 0
                if isinstance(set_cells, list):
                    for cell_item in set_cells:
                        if not isinstance(cell_item, dict):
                            continue
                        cell_ref = str(cell_item.get("cell", "")).strip().upper()
                        if not cell_ref:
                            continue
                        ws[cell_ref] = cell_item.get("value")
                        set_count += 1

                if isinstance(append_rows, list):
                    for row in append_rows:
                        if not isinstance(row, list):
                            continue
                        ws.append(row)
                        append_count += 1

                if isinstance(append_dict_rows, list):
                    for row_dict in append_dict_rows:
                        if not isinstance(row_dict, dict) or not row_dict:
                            continue
                        header_map: Dict[str, int] = {}
                        for col in range(1, ws.max_column + 1):
                            header_val = ws.cell(row=1, column=col).value
                            key = str(header_val).strip() if header_val is not None else ""
                            if key:
                                header_map[key] = col
                        if not header_map:
                            header_keys = list(row_dict.keys())
                            for idx, key in enumerate(header_keys, start=1):
                                ws.cell(row=1, column=idx, value=key)
                                header_map[str(key)] = idx
                        row_idx = ws.max_row + 1
                        for key, value in row_dict.items():
                            norm_key = str(key).strip()
                            if not norm_key:
                                continue
                            if norm_key not in header_map:
                                next_col = ws.max_column + 1
                                ws.cell(row=1, column=next_col, value=norm_key)
                                header_map[norm_key] = next_col
                            ws.cell(row=row_idx, column=header_map[norm_key], value=value)
                        append_count += 1

                if set_count == 0 and append_count == 0:
                    raise ValueError("xlsx_edit 未包含有效 set_cells/append_rows/append_dict_rows")

                target.parent.mkdir(parents=True, exist_ok=True)
                workbook.save(target)
                _append_changeset(
                    changeset,
                    rel,
                    action,
                    f"sheet={ws.title}, set_cells={set_count}, append_rows={append_count}",
                )
                logs.append(f"已更新 Excel: {rel}")
                continue

            errors.append(f"未知动作: {action}")
        except Exception as exc:
            errors.append(f"{action}({rel}) 失败: {exc}")

    return logs, changeset, errors


def file_edit_gateway_node(state: AppState) -> AppState:
    payload = state.get("payload", {})
    workspace_root = str(payload.get("workspace_root", "") or payload.get("workspace_dir", "")).strip()
    actions = payload.get("actions", [])
    operation_id = str(payload.get("operation_id", "")).strip() or "op-auto"
    policy = payload.get("policy", {}) if isinstance(payload.get("policy", {}), dict) else {}
    route_decision = state.get("route_decision", {}) if isinstance(state.get("route_decision", {}), dict) else {}
    requires_confirmation = bool(
        policy.get(
            "requires_confirmation",
            route_decision.get("requires_confirmation", False),
        )
    )
    user_confirmed = bool(policy.get("confirmed", False))

    errors = list(state.get("errors", []))
    warnings: List[str] = []

    if not workspace_root:
        err = "file_edit 任务缺少 workspace_root/workspace_dir"
        errors.append(err)
        return {
            "file_tool_result": {"status": "failed", "errors": [err]},
            "result": {"type": "file_edit", "status": "failed", "errors": errors},
            "errors": errors,
            "task_progress": state.get("task_progress", []) + [{"step": "file_edit_gateway", "tool_name": "FileToolGateway"}],
        }

    root = Path(workspace_root).expanduser().resolve()
    if not root.exists() or not root.is_dir():
        err = "workspace_root 不是有效目录"
        errors.append(err)
        return {
            "file_tool_result": {"status": "failed", "errors": [err]},
            "result": {"type": "file_edit", "status": "failed", "errors": errors},
            "errors": errors,
            "task_progress": state.get("task_progress", []) + [{"step": "file_edit_gateway", "tool_name": "FileToolGateway"}],
        }

    safe_actions = [item for item in actions if isinstance(item, dict)] if isinstance(actions, list) else []
    high_risk = any(str(item.get("action", "")).strip() in HIGH_RISK_ACTIONS for item in safe_actions)
    if high_risk and requires_confirmation and not user_confirmed:
        warnings.append("检测到高风险写操作，需先确认 policy.confirmed=true 后再执行。")
        result = {
            "type": "file_edit",
            "status": "pending_confirmation",
            "operation_id": operation_id,
            "warnings": warnings,
            "errors": errors,
            "changeset": [],
        }
        return {
            "file_tool_result": result,
            "result": result,
            "task_progress": state.get("task_progress", []) + [{"step": "file_edit_guard", "tool_name": "PolicyGuard"}],
        }

    logs, changeset, action_errors = _execute_actions(root, safe_actions)
    errors.extend(action_errors)
    status = "completed" if not action_errors else "partial_failed"
    result = {
        "type": "file_edit",
        "status": status,
        "operation_id": operation_id,
        "workspace_root": str(root),
        "logs": logs,
        "changeset": changeset,
        "warnings": warnings,
        "errors": errors,
    }
    return {
        "file_tool_result": result,
        "result": result,
        "errors": errors,
        "task_progress": state.get("task_progress", [])
        + [
            {"step": "file_edit_gateway", "tool_name": "FileToolGateway"},
        ],
    }
