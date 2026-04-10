"""
模板扫描器 - 自动扫描模板文件并生成配置文件
"""
import sys
from pathlib import Path
from typing import Dict, Any, List
import json
import re

# 添加项目根目录到Python路径
repo_root = Path(__file__).resolve().parents[2]
if str(repo_root) not in sys.path:
    sys.path.insert(0, str(repo_root))

import xlrd
from openpyxl import Workbook

from agent.parser.router import FileRouter
from agent.parser.schema import ParsedDocument

class TemplateScanner:
    """
    模板扫描器
    - 扫描templates目录中的所有文件
    - 解析模板文件，提取占位符信息
    - 生成templates_config.json配置文件
    """
    
    def __init__(self, templates_dir: str = "data/templates"):
        self.templates_dir = Path(templates_dir)
        self.router = FileRouter()
        self.placeholder_patterns = [
            r'\{\{([\w\u4e00-\u9fa5]+)\}\}',  # {{占位符}}
            r'\[([\w\u4e00-\u9fa5]+)\]',        # [占位符]
            r'__([\w\u4e00-\u9fa5]+)__',         # __占位符__
        ]
    
    def convert_xls_to_xlsx(self, xls_path: str, output_path: str = None) -> str:
        """
        将 .xls 文件转换为 .xlsx 文件
        
        Args:
            xls_path: .xls 文件路径
            output_path: 输出 .xlsx 文件路径，默认为同目录下同名 .xlsx 文件
            
        Returns:
            str: 转换后的 .xlsx 文件路径
        """
        try:
            xls_path = Path(xls_path)
            
            if output_path is None:
                output_path = xls_path.with_suffix('.xlsx')
            else:
                output_path = Path(output_path)
            
            # 读取 .xls 文件
            wb_xls = xlrd.open_workbook(str(xls_path))
            # 创建新的 .xlsx 文件
            wb_xlsx = Workbook()
            
            # 复制工作表和数据
            for sheet_name in wb_xls.sheet_names():
                ws_xls = wb_xls.sheet_by_name(sheet_name)
                ws_xlsx = wb_xlsx.create_sheet(title=sheet_name)
                
                # 复制数据
                for row in range(ws_xls.nrows):
                    for col in range(ws_xls.ncols):
                        ws_xlsx.cell(row=row+1, column=col+1).value = ws_xls.cell_value(row, col)
            
            # 删除默认的工作表
            if 'Sheet' in wb_xlsx.sheetnames:
                wb_xlsx.remove(wb_xlsx['Sheet'])
            
            # 保存为 .xlsx 文件
            wb_xlsx.save(str(output_path))
            
            return str(output_path)
        except Exception as e:
            print(f"转换 .xls 文件失败: {str(e)}")
            raise
    
    def batch_convert_xls_to_xlsx(self) -> List[str]:
        """
        批量转换目录中的 .xls 文件为 .xlsx 文件
        
        Returns:
            List[str]: 转换成功的文件路径列表
        """
        converted_files = []
        
        for xls_file in self.templates_dir.glob('*.xls'):
            try:
                xlsx_file = xls_file.with_suffix('.xlsx')
                print(f"转换 {xls_file.name} 为 {xlsx_file.name}")
                converted_path = self.convert_xls_to_xlsx(str(xls_file), str(xlsx_file))
                converted_files.append(converted_path)
                print(f"转换成功: {converted_path}")
            except Exception as e:
                print(f"转换失败: {xls_file.name}, 错误: {str(e)}")
        
        return converted_files
    
    def scan_all_templates(self) -> Dict[str, Any]:
        """
        扫描所有模板文件
        """
        config = {}
        
        # 先转换所有 .xls 文件为 .xlsx 文件
        self.batch_convert_xls_to_xlsx()
        
        # 支持的文件类型
        supported_extensions = [".docx", ".xlsx", ".pdf"]
        
        # 扫描目录
        for template_file in self.templates_dir.iterdir():
            # 跳过临时文件和隐藏文件
            if template_file.name.startswith('~$') or template_file.name.startswith('.'):
                continue
                
            if template_file.suffix.lower() not in supported_extensions:
                continue
            
            template_name = template_file.name
            template_id = template_name.split('.')[0].lower().replace(' ', '_')
            
            try:
                # 解析模板文件
                parsed_doc = self.router.parse_file(str(template_file))
                
                # 提取占位符
                placeholders = self._extract_placeholders(parsed_doc)
                
                # 生成字段映射
                field_mapping = self._generate_field_mapping(placeholders)
                
                # 生成配置
                template_config = self._generate_template_config(
                    template_file, parsed_doc, placeholders, field_mapping
                )
                
                config[template_id] = template_config
                
            except Exception as e:
                print(f"Error scanning template {template_name}: {e}")
                continue
        
        return config
    


    
    def _extract_placeholders(self, doc: ParsedDocument) -> List[str]:
        """
        从解析结果中提取占位符
        """
        placeholders = []
        
        # 从段落中提取
        for section in doc.sections:
            found = self._find_placeholders_in_text(section.content)
            placeholders.extend(found)
        
        # 从表格中提取
        for table in doc.tables:
            for row in table.rows:
                for cell in row:
                    if isinstance(cell, str):
                        found = self._find_placeholders_in_text(cell)
                        placeholders.extend(found)
        
        return list(set(placeholders))
    
    def _find_placeholders_in_text(self, text: str) -> List[str]:
        """
        在文本中查找占位符
        """
        placeholders = []
        
        for pattern in self.placeholder_patterns:
            matches = re.findall(pattern, text)
            placeholders.extend(matches)
        
        return placeholders
    
    def _generate_field_mapping(self, placeholders: List[str]) -> Dict[str, str]:
        """
        生成字段映射
        """
        field_mapping = {}
        
        # 标准映射
        standard_mapping = {
            "经办人姓名": "student_name",
            "经办人联系方式": "contact",
            "活动时间": "activity_time",
            "活动地点": "location",
            "参与人员": "participants",
            "活动主要内容": "description",
            "报销内容及金额": "expense_detail",
            "发票序号": "invoice_no",
            "发票金额": "amount",
            "发票日期": "invoice_date",
            "发票内容": "content",
            "具体活动名称": "activity_name",
            "活动举办日期": "activity_date",
            "归属（学生组织）": "org",
            "经办同学": "student_name",
            "学号": "student_id"
        }
        
        for placeholder in placeholders:
            # 优先使用标准映射
            if placeholder in standard_mapping:
                field_mapping[standard_mapping[placeholder]] = placeholder
            else:
                # 自动生成字段名
                field_name = placeholder.lower().replace(' ', '_').replace('（', '_').replace('）', '')
                field_mapping[field_name] = placeholder
        
        return field_mapping
    
    def _generate_template_config(
        self,
        template_file: Path,
        doc: ParsedDocument,
        placeholders: List[str],
        field_mapping: Dict[str, str]
    ) -> Dict[str, Any]:
        """
        生成单个模板的配置
        """
        template_type = self._detect_template_type(template_file, doc)
        
        metadata = {
            "content_type": doc.content_type,
            "sections_count": len(doc.sections),
            "tables_count": len(doc.tables),
            "placeholders_count": len(placeholders)
        }
        
        # 从文件名中提取信息
        parent_folders = []
        if '_' in template_file.name:
            parts = template_file.name.split('_', 1)
            if len(parts) == 2:
                folder_prefix = parts[0]
                parent_folders = [folder_prefix]
        
        metadata.update({
            "parent_folders": parent_folders,
            "folder_depth": len(parent_folders)
        })
        
        return {
            "filename": template_file.name,
            "type": template_type,
            "placeholders": placeholders,
            "field_mapping": field_mapping,
            "output_filename_pattern": f"{template_file.stem}_{{timestamp}}{template_file.suffix}",
            "metadata": metadata
        }
    
    def _detect_template_type(self, template_file: Path, doc: ParsedDocument) -> str:
        """
        检测模板类型
        """
        if template_file.suffix == ".docx":
            return "word"
        elif template_file.suffix == ".xlsx":
            return "excel"
        elif template_file.suffix == ".pdf":
            return "pdf"
        return "unknown"
    
    def generate_config_file(self, output_path: str = "data/templates/templates_config.json"):
        """
        生成配置文件
        """
        config = self.scan_all_templates()
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        
        print(f"Generated template config file: {output_path}")
        print(f"Total templates: {len(config)}")
        
        return config

if __name__ == "__main__":
    scanner = TemplateScanner()
    scanner.generate_config_file()
