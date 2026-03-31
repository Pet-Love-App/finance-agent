"""
agent/parser/parsers/excel_parser.py

Excel 解析器 —— BOO-63 重点。
必须支持：
- 列名别名解析（写入 warnings + 置信度）
- 空行/空列跳过（写 warnings）
- 合并单元格拆包（warnings + 策略记录）
- 混合类型容错（标准化 + 可回溯原值）
- 多表识别（同一 Sheet 多个表块 → 拆分 table_id）
- 坐标映射：csv(row,col) → Excel A1
- 输出 tables/<table_id>.csv + tables/<table_id>.format.json
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Optional

from agent.parser.base import BaseParser
from agent.parser.schema import (
    Error, Loc, LocType, ParsedDocument, TableBlock, TableMeta, Warning,
)
from agent.parser.utils.hash_utils import file_sha1
from agent.parser.utils.file_utils import excel_a1, excel_range

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    supported_extensions = [".xlsx", ".xls"]

    def __init__(
        self,
        kb_name: str = "",
        max_empty_rows: int = 5,     # 连续空行超过此值 → 认为表块结束
        max_rows: int = 50000,       # 单 sheet 行数上限（防止 OOM）
    ):
        self.kb_name = kb_name
        self.max_empty_rows = max_empty_rows
        self.max_rows = max_rows

    def parse(self, file_path: str, **kwargs) -> ParsedDocument:
        import openpyxl

        doc_id = file_sha1(file_path, prefix=self.kb_name)
        fpath = Path(file_path)
        warnings: list[Warning] = []
        errors: list[Error] = []
        all_tables: list[TableBlock] = []

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as exc:
            return ParsedDocument(
                doc_id=doc_id,
                source=self._source_dict(fpath),
                content_type="excel",
                errors=[Error(
                    error_code="EXCEL_OPEN_FAILED",
                    message=str(exc),
                    loc=Loc(type=LocType.SHEET.value, value=""),
                )],
            )

        sheet_names = wb.sheetnames
        global_table_idx = 0

        for sheet_name in sheet_names:
            ws = wb[sheet_name]

            # 检查 sheet 是否为空
            if ws.max_row is None or ws.max_row == 0:
                warnings.append(Warning(
                    code="EMPTY_SHEET",
                    message=f"Sheet '{sheet_name}' is empty.",
                    loc=Loc(type=LocType.SHEET.value, value=sheet_name),
                ))
                continue

            if ws.max_row > self.max_rows:
                warnings.append(Warning(
                    code="SHEET_TRUNCATED",
                    message=(
                        f"Sheet '{sheet_name}' has {ws.max_row} rows, "
                        f"truncated to {self.max_rows}."
                    ),
                    loc=Loc(type=LocType.SHEET.value, value=sheet_name),
                ))

            # ---- Step 1: 读取所有单元格为原始值矩阵 ----
            raw_grid = self._read_raw_grid(ws, warnings, sheet_name)

            # ---- Step 2: 处理合并单元格 ----
            raw_grid = self._expand_merged_cells(ws, raw_grid, warnings, sheet_name)

            # ---- Step 3: 检测表块（可能一个 Sheet 多个表）----
            table_blocks = self._detect_table_blocks(raw_grid, warnings, sheet_name)

            if not table_blocks:
                # 整个 sheet 当一个表处理
                table_blocks = [(0, 0, len(raw_grid) - 1, len(raw_grid[0]) - 1 if raw_grid else 0)]

            # ---- Step 4: 逐个表块提取 ----
            for block_idx, (r1, c1, r2, c2) in enumerate(table_blocks):
                global_table_idx += 1
                table_id = f"t{global_table_idx}"
                try:
                    tb = self._extract_table_block(
                        raw_grid, r1, c1, r2, c2,
                        table_id, sheet_name, warnings
                    )
                    all_tables.append(tb)
                except Exception as exc:
                    errors.append(Error(
                        error_code="TABLE_BLOCK_EXTRACT_FAILED",
                        message=f"Sheet '{sheet_name}' block {block_idx+1}: {exc}",
                        loc=Loc(
                            type=LocType.RANGE.value,
                            value=excel_range(r1, c1, r2, c2),
                            extra={"sheet": sheet_name},
                        ),
                    ))

        # ---- 收集格式元信息 ----
        format_meta = self._collect_format_meta(wb, fpath.name)

        wb.close()

        return ParsedDocument(
            doc_id=doc_id,
            source=self._source_dict(fpath),
            title=fpath.stem,
            content_type="excel",
            tables=all_tables,
            warnings=warnings,
            errors=errors,
            metadata={
                "sheet_names": sheet_names,
                "total_tables": len(all_tables),
                "format_meta": format_meta,
            },
        )

    # ==================================================================
    # Step 1: 读取原始网格
    # ==================================================================
    def _read_raw_grid(
        self, ws, warnings: list[Warning], sheet_name: str
    ) -> list[list[Any]]:
        """将 worksheet 读为二维列表（保留原始值）"""
        grid: list[list[Any]] = []
        max_row = min(ws.max_row or 0, self.max_rows)
        max_col = ws.max_column or 0

        empty_row_count = 0
        for row_idx in range(1, max_row + 1):
            row_values = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_values.append(cell.value)
            # 空行检测
            if all(v is None or str(v).strip() == "" for v in row_values):
                empty_row_count += 1
            else:
                # 如果之前有空行被跳过，记录
                if empty_row_count > 0:
                    if empty_row_count >= 3:
                        warnings.append(Warning(
                            code="EMPTY_ROWS_SKIPPED",
                            message=(
                                f"Sheet '{sheet_name}': {empty_row_count} consecutive "
                                f"empty rows before row {row_idx}."
                            ),
                            loc=Loc(
                                type=LocType.CELL.value,
                                value=excel_a1(row_idx - empty_row_count - 1, 0),
                                extra={"sheet": sheet_name, "count": empty_row_count},
                            ),
                        ))
                empty_row_count = 0
            grid.append(row_values)

        return grid

    # ==================================================================
    # Step 2: 合并单元格展开
    # ==================================================================
    def _expand_merged_cells(
        self, ws, grid: list[list[Any]],
        warnings: list[Warning], sheet_name: str,
    ) -> list[list[Any]]:
        """展开合并单元格 —— 将合并区域的值填充到所有被合并的格子"""
        merged_ranges = list(ws.merged_cells.ranges)
        if not merged_ranges:
            return grid

        expanded_count = 0
        for mr in merged_ranges:
            min_r, min_c = mr.min_row - 1, mr.min_col - 1
            max_r, max_c = mr.max_row - 1, mr.max_col - 1
            # 取合并区域的左上角值
            top_left_val = grid[min_r][min_c] if min_r < len(grid) and min_c < len(grid[min_r]) else None
            for r in range(min_r, min(max_r + 1, len(grid))):
                for c in range(min_c, min(max_c + 1, len(grid[r]) if r < len(grid) else 0)):
                    if r == min_r and c == min_c:
                        continue
                    if r < len(grid) and c < len(grid[r]):
                        grid[r][c] = top_left_val
                        expanded_count += 1

        if expanded_count > 0:
            warnings.append(Warning(
                code="MERGED_CELLS_EXPANDED",
                message=(
                    f"Sheet '{sheet_name}': {len(merged_ranges)} merged region(s), "
                    f"{expanded_count} cell(s) filled with top-left value. "
                    f"Strategy: fill-from-top-left."
                ),
                loc=Loc(
                    type=LocType.SHEET.value,
                    value=sheet_name,
                    extra={
                        "merged_ranges": [str(mr) for mr in merged_ranges],
                        "strategy": "fill-from-top-left",
                    },
                ),
            ))

        return grid

    # ==================================================================
    # Step 3: 多表块检测
    # ==================================================================
    def _detect_table_blocks(
        self, grid: list[list[Any]],
        warnings: list[Warning], sheet_name: str,
    ) -> list[tuple[int, int, int, int]]:
        """
        检测同一 Sheet 中的多个表块。
        启发式：连续 N 个全空行视为表块分隔。
        返回 (r1, c1, r2, c2) 列表。
        """
        if not grid:
            return []

        blocks: list[tuple[int, int, int, int]] = []
        block_start: Optional[int] = None
        empty_count = 0
        max_col = max(len(row) for row in grid) if grid else 0

        for ri, row in enumerate(grid):
            is_empty = all(v is None or str(v).strip() == "" for v in row)
            if is_empty:
                empty_count += 1
                if empty_count >= self.max_empty_rows and block_start is not None:
                    # 结束当前块
                    block_end = ri - empty_count
                    if block_end >= block_start:
                        # 裁剪空列
                        c1, c2 = self._trim_empty_cols(grid, block_start, block_end, max_col)
                        blocks.append((block_start, c1, block_end, c2))
                    block_start = None
            else:
                if block_start is None:
                    block_start = ri
                empty_count = 0

        # 最后一个块
        if block_start is not None:
            block_end = len(grid) - 1
            # 去掉末尾空行
            while block_end > block_start and all(
                v is None or str(v).strip() == ""
                for v in grid[block_end]
            ):
                block_end -= 1
            c1, c2 = self._trim_empty_cols(grid, block_start, block_end, max_col)
            blocks.append((block_start, c1, block_end, c2))

        if len(blocks) > 1:
            warnings.append(Warning(
                code="MULTIPLE_TABLES_DETECTED",
                message=(
                    f"Sheet '{sheet_name}': {len(blocks)} table block(s) detected "
                    f"(separated by {self.max_empty_rows}+ empty rows)."
                ),
                loc=Loc(type=LocType.SHEET.value, value=sheet_name,
                        extra={"block_ranges": [
                            excel_range(r1, c1, r2, c2) for r1, c1, r2, c2 in blocks
                        ]}),
            ))

        return blocks

    @staticmethod
    def _trim_empty_cols(
        grid: list[list[Any]], r1: int, r2: int, max_col: int
    ) -> tuple[int, int]:
        """找到非空列的起止范围"""
        min_c, max_c = max_col, 0
        for ri in range(r1, r2 + 1):
            for ci in range(len(grid[ri])):
                val = grid[ri][ci]
                if val is not None and str(val).strip():
                    min_c = min(min_c, ci)
                    max_c = max(max_c, ci)
        if min_c > max_c:
            return 0, max(max_col - 1, 0)
        return min_c, max_c

    # ==================================================================
    # Step 4: 提取单个表块
    # ==================================================================
    def _extract_table_block(
        self,
        grid: list[list[Any]],
        r1: int, c1: int, r2: int, c2: int,
        table_id: str, sheet_name: str,
        warnings: list[Warning],
    ) -> TableBlock:
        """从 grid 中提取 [r1:r2+1, c1:c2+1] 范围的表块"""

        # 截取子矩阵
        sub = []
        for ri in range(r1, r2 + 1):
            row = grid[ri] if ri < len(grid) else []
            sub_row = []
            for ci in range(c1, c2 + 1):
                val = row[ci] if ci < len(row) else None
                sub_row.append(val)
            sub.append(sub_row)

        if not sub:
            raise ValueError("Empty table block")

        # 表头（第一行）
        raw_headers = sub[0]
        headers = []
        for ci, h in enumerate(raw_headers):
            if h is None or str(h).strip() == "":
                col_name = f"col_{ci + 1}"
                warnings.append(Warning(
                    code="EMPTY_HEADER",
                    message=f"Table {table_id}: column {ci+1} has empty header, named '{col_name}'.",
                    loc=Loc(
                        type=LocType.CELL.value,
                        value=excel_a1(r1, c1 + ci),
                        extra={"sheet": sheet_name},
                    ),
                ))
                headers.append(col_name)
            else:
                headers.append(str(h).strip())

        # 数据行
        data_rows = sub[1:]

        # 混合类型检测 + 标准化
        data_rows, type_warnings = self._normalize_values(
            data_rows, headers, r1 + 1, c1, sheet_name, table_id
        )
        warnings.extend(type_warnings)

        # 坐标映射：csv (data_row_idx, col_idx) → Excel A1
        coord_map = {}
        for ri in range(len(data_rows)):
            for ci in range(len(headers)):
                # csv 行号 ri 对应 Excel 行 r1 + 1 + ri (header 在 r1)
                excel_row = r1 + 1 + ri
                excel_col = c1 + ci
                coord_map[f"{ri},{ci}"] = excel_a1(excel_row, excel_col)

        # 合并单元格信息（已在 grid 级别展开，此处记录原始范围）
        source_range = excel_range(r1, c1, r2, c2)

        return TableBlock(
            meta=TableMeta(
                table_id=table_id,
                source_sheet=sheet_name,
                source_range=source_range,
                header_rows=[1],
                row_count=len(data_rows),
                col_count=len(headers),
                coord_map=coord_map,
            ),
            headers=headers,
            rows=data_rows,
        )

    # ==================================================================
    # 混合类型容错
    # ==================================================================
    def _normalize_values(
        self,
        rows: list[list[Any]],
        headers: list[str],
        data_start_row: int,  # Excel 行号 (0-based)
        col_offset: int,      # Excel 列偏移 (0-based)
        sheet_name: str,
        table_id: str,
    ) -> tuple[list[list[Any]], list[Warning]]:
        """
        逐列检测类型，对混合类型进行标准化。
        - 金额列出现字符串 → 尝试提取数值 + warning
        - 千分位 / 货币符号 → 清理
        """
        warngs: list[Warning] = []
        if not rows or not rows[0]:
            return rows, warngs

        n_cols = len(headers)
        # 列类型推断
        col_types = self._infer_column_types(rows, n_cols)

        for ci in range(n_cols):
            expected = col_types[ci]
            if expected == "numeric":
                for ri in range(len(rows)):
                    val = rows[ri][ci] if ci < len(rows[ri]) else None
                    if val is None:
                        continue
                    if isinstance(val, (int, float)):
                        continue
                    # 尝试清理并转为数值
                    cleaned, ok = self._try_parse_numeric(str(val))
                    if ok:
                        original = str(val)
                        rows[ri][ci] = cleaned
                        if original != str(cleaned):
                            warngs.append(Warning(
                                code="MIXED_TYPE_NORMALIZED",
                                message=(
                                    f"Table {table_id}, col '{headers[ci]}' row {ri+1}: "
                                    f"'{original}' → {cleaned}"
                                ),
                                loc=Loc(
                                    type=LocType.CELL.value,
                                    value=excel_a1(data_start_row + ri, col_offset + ci),
                                    extra={"sheet": sheet_name, "original": original},
                                ),
                            ))
                    else:
                        # 无法转换，保留原值 + warning
                        warngs.append(Warning(
                            code="MIXED_TYPE_UNCONVERTED",
                            message=(
                                f"Table {table_id}, col '{headers[ci]}' row {ri+1}: "
                                f"expected numeric but got '{val}'"
                            ),
                            loc=Loc(
                                type=LocType.CELL.value,
                                value=excel_a1(data_start_row + ri, col_offset + ci),
                                extra={"sheet": sheet_name},
                            ),
                        ))

        return rows, warngs

    @staticmethod
    def _infer_column_types(rows: list[list[Any]], n_cols: int) -> list[str]:
        """简单列类型推断：numeric / text / mixed"""
        types = []
        for ci in range(n_cols):
            num_count = 0
            text_count = 0
            for row in rows:
                val = row[ci] if ci < len(row) else None
                if val is None or str(val).strip() == "":
                    continue
                if isinstance(val, (int, float)):
                    num_count += 1
                else:
                    # 尝试解析
                    s = str(val).strip()
                    s_clean = re.sub(r'[¥￥$,，\s]', '', s)
                    try:
                        float(s_clean)
                        num_count += 1
                    except ValueError:
                        text_count += 1
            total = num_count + text_count
            if total == 0:
                types.append("text")
            elif num_count / total > 0.7:
                types.append("numeric")
            else:
                types.append("text")
        return types

    @staticmethod
    def _try_parse_numeric(s: str) -> tuple[Any, bool]:
        """尝试将字符串解析为数值"""
        s = s.strip()
        # 去除货币符号、千分位
        cleaned = re.sub(r'[¥￥$€£]', '', s)
        cleaned = cleaned.replace(',', '').replace('，', '').strip()
        # 处理 "约xxx元" "xxx元" 等
        m = re.search(r'[\d.]+', cleaned)
        if m:
            try:
                val = float(m.group())
                if val == int(val):
                    return int(val), True
                return val, True
            except ValueError:
                pass
        return s, False

    # ==================================================================
    # 格式元信息
    # ==================================================================
    def _collect_format_meta(self, wb, workbook_name: str) -> dict:
        """收集影响语义解释的格式信息"""
        meta: dict[str, Any] = {"workbook": workbook_name, "sheets": {}}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_meta: dict[str, Any] = {}

            # 合并单元格
            merged = [str(mr) for mr in ws.merged_cells.ranges]
            if merged:
                sheet_meta["merged_cells"] = merged

            # 冻结窗格
            if ws.freeze_panes:
                sheet_meta["freeze_panes"] = str(ws.freeze_panes)

            # 数字格式（抽样前 10 行）
            number_formats: dict[str, set] = {}
            for row_idx, row in enumerate(ws.iter_rows(max_row=min(10, ws.max_row or 0))):
                for cell in row:
                    if cell.number_format and cell.number_format != "General":
                        col_letter = cell.column_letter
                        number_formats.setdefault(col_letter, set()).add(cell.number_format)
            if number_formats:
                sheet_meta["number_formats"] = {
                    k: list(v) for k, v in number_formats.items()
                }

            if sheet_meta:
                meta["sheets"][sheet_name] = sheet_meta

        return meta

    # ==================================================================
    # 辅助
    # ==================================================================
    def _source_dict(self, fpath: Path) -> dict:
        return {
            "file_name": fpath.name,
            "file_type": fpath.suffix.lstrip("."),
            "file_size_bytes": fpath.stat().st_size,
            "file_path": str(fpath),
        }